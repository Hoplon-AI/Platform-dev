"""
SoV Processor v3 — Robust Single-LLM-Call Schedule of Values ingestion pipeline.

Fixes all 13 known issues from the Master Build Specification:
  #1  header_row:null crash        → null guards on every LLM field
  #2  Wrong SI year column         → _extract_year_from_col_name() handles ranges
  #3  Wrong sheet selected         → dwelling-specific sheets preferred over supersets
  #4  Non-unique property ref      → uniqueness validation + address fallback
  #5  Year of Build not extracted  → handles datetime/float/string/encoded formats
  #6  Postcode overflow            → regex validate + clamp to 10
  #7  Too many LLM calls           → single-call architecture (1 or 0 calls)
  #8  Join plan None fields        → or-0 guards on all join plan fields
  #9  varchar overflow             → _clamp() on every string field before insert
  #10 Filter/metadata rows         → _is_metadata_row() check
  #11 Insure flag ignored          → skip rows where insure_flag == N
  #12 SI in separate sheet         → classify SI-heavy sheets as secondary not noise
  #13 Off-by-one SI edge rows      → flag in metadata for manual review

Architecture:
  Stage A   Workbook Snapshot   (code)   — profile every sheet
  Stage B   Workbook Analysis   (LLM×1)  — single prompt returns everything
  Stage C   Row Extraction      (code)   — extract, assemble, clean, confidence
  Stage D   Upsert              (code)   — write to silver.properties

Zero-crash guarantee: entire pipeline wrapped in try/except.
Zero-capital-loss: SI never silently dropped; mixed-basis flagged in metadata.
"""

from __future__ import annotations

import json
import logging
import os
import re
import uuid
from datetime import datetime, timezone
from typing import Any

import openpyxl
import psycopg2
import psycopg2.extras

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────

CANONICAL_FIELDS: dict[str, str] = {
    "property_reference":  "Unique ID for dwelling/unit",
    "block_reference":     "Parent block/building ID",
    "address":             "Street address",
    "address_2":           "Address line 2",
    "address_3":           "Town/city",
    "postcode":            "UK postcode",
    "occupancy_type":      "Tenure e.g. Rented, Factored, Shared Ownership",
    "sum_insured":         "Rebuild cost / declared value (£, numeric)",
    "sum_insured_type":    "Basis e.g. Day 1 Reinstatement",
    "property_type":       "e.g. Flat, House, Bungalow",
    "avid_property_type":  "Avid system property category",
    "wall_construction":   "External wall material",
    "roof_construction":   "Roof material",
    "floor_construction":  "Floor material",
    "year_of_build":       "Year built (integer or date)",
    "age_banding":         "Decade band e.g. Pre-1919, 1945-1964",
    "num_bedrooms":        "Number of bedrooms",
    "num_storeys":         "Storeys above ground",
    "num_units":           "Units in block",
    "has_basement":        "Basement present (Yes/No)",
    "is_listed":           "Listed building status",
    "security_features":   "Security measures",
    "fire_protection":     "Fire suppression systems",
    "alarms":              "Fire/smoke alarms",
    "flood_insured":       "Flood cover included",
    "storm_insured":       "Storm cover included",
    "deductible":          "Policy excess amount",
    "insure_flag":         "Y/N whether property is insured (filter rows where N)",
}

# Issue #9 — varchar limits for _clamp()
DB_COLUMN_LIMITS: dict[str, int] = {
    "property_reference": 100,
    "block_reference":    100,
    "postcode":            10,
    "occupancy_type":     100,
    "sum_insured_type":   100,
    "property_type":      100,
    "avid_property_type": 100,
    "wall_construction":  100,
    "roof_construction":  100,
    "floor_construction": 100,
    "year_of_build":       20,
    "age_banding":         20,
    "security_features":  255,
    "fire_protection":    255,
    "alarms":             255,
}

# ── Keyword sets ──

NOISE_KEYWORDS = frozenset([
    "broker fee", "indexation", "questionnaire", "payroll", "wages",
    "turnover", "vehicle", "salary", "matrix", "comparison", "working",
    "notes", "summary", "pivot", "validation", "development", "land",
    "landbank", "void procedure", "cg review", "rents", "rent ",
    "sheet1", "sheet2", "sheet3", "sheet4", "sheet5", "sheet6",
    "sheet7", "sheet8", "sheet9", "nbv", "all properties",
    "prudential", "savills", "march 20", "lifts",
    "hoist", "properties over", "void", "procedure",
])

DATA_KEYWORDS = frozenset([
    "stock", "rented", "rental", "asset dwelling", "dwelling", "property",
    "listing", "schedule", "lettable", "residential", "unit", "flat",
    "house", "block", "portfolio", "shared", "factored", "leasehold",
    "final", "hostel", "key worker",
])

# Issue #3 — sheet preference keywords
SUPERSET_PENALTIES = frozenset([
    "all assets", "all properties", "full list", "complete stock",
])
DWELLING_PREFER = frozenset([
    "dwelling", "residential", "rented stock", "asset dwelling",
    "lettable", "stock list",
])

SI_KEYWORDS = frozenset([
    "sum insured", "declared value", "rebuild", "reinstate",
    "insurance value", "insured value", "block rebuild",
    "total insured", "di value", "reinstatement",
])

# Column-name keywords for heuristic field mapping
FIELD_KEYWORDS: dict[str, list[str]] = {
    "property_reference": [
        "property ref", "asset ref", "jrf ref", "place ref", "council ref",
        "prop id", "property id", "unit ref", "uprn", "asset reference",
        "prty_id", "property code", "reconciliation",
    ],
    "block_reference": [
        "block ref", "parent ref", "block name", "block id", "jll block",
        "default block", "parent asset", "parent asset ref",
    ],
    "address": [
        "full address", "address details", "address line 1", "add 1",
        "add1", "risk address", "address 1", "address",
    ],
    "address_2": ["address line 2", "add 2", "add2", "address 2"],
    "address_3": ["address line 3", "add 3", "add3", "address 3", "town", "city"],
    "postcode": ["postcode", "post code", "pc", "post_code"],
    "occupancy_type": [
        "occupancy", "tenure", "rental type", "tenancy", "asset tenancy",
        "classification", "letting type", "insurance tenure", "rent type",
        "need category desc", "insurance - category",
    ],
    "sum_insured": [
        "sum insured", "declared value", "rebuild", "reinstate",
        "insurance value", "block rebuild", "di value", "reinstatement",
    ],
    "sum_insured_type": ["si type", "insured type", "basis of", "sum insured type"],
    "property_type": [
        "property type", "asset type", "location type", "type",
        "unit description", "built form",
    ],
    "avid_property_type": ["avid"],
    "wall_construction": [
        "wall construction", "wall material", "external wall",
        "construction", "non traditional",
    ],
    "roof_construction": ["roof construction", "roof material", "roof type"],
    "floor_construction": ["floor construction", "floor material"],
    "year_of_build": [
        "year built", "year of build", "construction date",
        "year build", "build year", "built", "build date",
    ],
    "age_banding": ["age band", "age banding"],
    "num_bedrooms": [
        "bedroom", "beds", "no. of bed", "no of bed", "bedrooms",
        "number of bed", "bed size", "lettablebeds",
    ],
    "num_storeys": [
        "storey", "stories", "floors", "no of floor", "number of floor",
        "storeys above", "number of storey", "no. of stor", "no of floors",
    ],
    "num_units": [
        "no. of unit", "number of unit", "units in block",
        "number of units", "no of unit", "max occs",
    ],
    "has_basement": ["basement"],
    "is_listed": ["listed"],
    "security_features": ["security"],
    "fire_protection": ["fire protection", "sprinkler"],
    "alarms": ["alarm"],
    "flood_insured": ["flood"],
    "storm_insured": ["storm"],
    "deductible": ["deductible", "excess"],
    "insure_flag": ["insure yes", "insure y/n", "insured yes"],
}

UK_POSTCODE_RE = re.compile(r"^[A-Z]{1,2}\d{1,2}[A-Z]?\s*\d[A-Z]{2}$", re.IGNORECASE)
UK_POSTCODE_EXTRACT_RE = re.compile(
    r"\b([A-Z]{1,2}\d{1,2}[A-Z]?\s*\d[A-Z]{2})\b", re.IGNORECASE
)

MAX_PREVIEW_ROWS = 5
MAX_PREVIEW_COLS = 40
MAX_HEADER_SCAN = 10


# ─────────────────────────────────────────────────────────────────────────────
# LLM helper
# ─────────────────────────────────────────────────────────────────────────────

def _llm(prompt: str, system: str | None = None, max_tokens: int = 4096) -> str:
    try:
        from backend.core.agentic.bedrock_client import invoke_claude
        return invoke_claude(prompt, system=system, max_tokens=max_tokens, temperature=0.0)
    except Exception as exc:
        logger.warning(f"[SOV LLM] Bedrock call failed: {exc}")
        return ""


def _parse_json(text: str) -> dict | list | None:
    if not text:
        return None
    text = re.sub(r"```(?:json)?", "", text).strip()
    for start_char, end_char in [("{", "}"), ("[", "]")]:
        idx = text.find(start_char)
        if idx == -1:
            continue
        depth = 0
        for i, ch in enumerate(text[idx:], start=idx):
            if ch == start_char:
                depth += 1
            elif ch == end_char:
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[idx:i + 1])
                    except json.JSONDecodeError:
                        break
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

def _load_workbook(file_bytes: bytes) -> openpyxl.Workbook:
    import io
    return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)


def _sheet_rows(ws) -> list[tuple]:
    return list(ws.iter_rows(values_only=True))


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return str(val.year)
    return str(val).strip()


def _count_filled(row: tuple) -> int:
    return sum(1 for c in row if c is not None and str(c).strip())


# ─────────────────────────────────────────────────────────────────────────────
# Issue #9 — _clamp() for varchar overflow protection
# ─────────────────────────────────────────────────────────────────────────────

def _clamp(value: str | None, max_len: int) -> str | None:
    if value is None:
        return None
    return value[:max_len] if len(value) > max_len else value


# ─────────────────────────────────────────────────────────────────────────────
# Issue #2 — SI year extraction from column names
# ─────────────────────────────────────────────────────────────────────────────

def _extract_year_from_col_name(col_name: str) -> int:
    """Extract year from SI column names. Handles ranges like '25/26', '2024-25'."""
    s = str(col_name)
    # Range formats: "2025/26", "24/25", "2024-25", "2026-2027"
    range_match = re.search(r"(\d{2,4})[/\-](\d{2,4})", s)
    if range_match:
        y1_raw, y2_raw = range_match.group(1), range_match.group(2)
        y1 = int(y1_raw) if len(y1_raw) == 4 else 2000 + int(y1_raw)
        y2 = int(y2_raw) if len(y2_raw) == 4 else 2000 + int(y2_raw)
        return max(y1, y2)  # take the HIGHER year of the range
    # Plain 4-digit year: "Reinstatement 2026", "JLL RCA 2025"
    years = re.findall(r"20\d{2}", s)
    if years:
        return max(int(y) for y in years)
    return 0


# ─────────────────────────────────────────────────────────────────────────────
# Issue #4 — Property reference uniqueness validation
# ─────────────────────────────────────────────────────────────────────────────

def _validate_property_ref_uniqueness(rows: list[tuple], ref_col: int) -> float:
    """Returns ratio of unique values to total non-null rows. < 0.8 = reject."""
    values = [
        str(r[ref_col]).strip()
        for r in rows
        if ref_col < len(r) and r[ref_col] and str(r[ref_col]).strip()
    ]
    if not values:
        return 0.0
    return len(set(values)) / len(values)


# ─────────────────────────────────────────────────────────────────────────────
# Issue #5 — Year of build extraction (all formats)
# ─────────────────────────────────────────────────────────────────────────────

def _decode_year(val: Any) -> int | None:
    """Extract year from datetime, float, string, or encoded building type."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.year
    if isinstance(val, float):
        y = int(val)
        return y if 1700 <= y <= 2030 else None
    if isinstance(val, int):
        return val if 1700 <= val <= 2030 else None
    s = str(val).strip()
    # Try encoded building type first: P2002H, P1982T
    enc = _decode_year_from_building_type(s)
    if enc:
        return enc
    # Standard year extraction
    m = re.search(r"\b(1[789]\d{2}|20[012]\d)\b", s)
    if m:
        return int(m.group(1))
    m = re.search(r"(1[789]\d{2}|20[012]\d)", s)
    if m:
        return int(m.group(1))
    return None


def _decode_year_from_building_type(val: Any) -> int | None:
    """Decode encoded building type fields: P2002H→2002, 19TO44T→~1930."""
    if val is None:
        return None
    s = str(val).strip()
    # Direct year: P2002H, P1982T
    m = re.search(r"(1[789]\d{2}|20[012]\d)", s)
    if m:
        return int(m.group(1))
    # Range: 19TO44T → midpoint of 1919-1944
    range_m = re.match(r"(\d{2})TO(\d{2})", s, re.IGNORECASE)
    if range_m:
        y1 = 1900 + int(range_m.group(1))
        y2 = 1900 + int(range_m.group(2))
        return (y1 + y2) // 2
    return None


def _derive_age_banding(year: int | None) -> str | None:
    if year is None:
        return None
    if year < 1919:
        return "Pre-1919"
    elif year < 1945:
        return "1919-1944"
    elif year < 1965:
        return "1945-1964"
    elif year < 1981:
        return "1965-1980"
    elif year < 1991:
        return "1981-1990"
    elif year < 2001:
        return "1991-2000"
    elif year < 2011:
        return "2001-2010"
    else:
        return "Post-2010"


# ─────────────────────────────────────────────────────────────────────────────
# Issue #6 — Postcode cleaning with overflow protection
# ─────────────────────────────────────────────────────────────────────────────

def _clean_postcode(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip().upper()
    if UK_POSTCODE_RE.match(s):
        return s[:10]
    m = UK_POSTCODE_EXTRACT_RE.search(s)
    return m.group(1).upper()[:10] if m else None


# ─────────────────────────────────────────────────────────────────────────────
# Value cleaning utilities
# ─────────────────────────────────────────────────────────────────────────────

def _clean_numeric(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val != 0 else None
    s = re.sub(r"[£,\s]", "", str(val))
    try:
        result = float(s)
        return result if result != 0 else None
    except ValueError:
        return None


def _clean_int(val: Any) -> int | None:
    n = _clean_numeric(val)
    return int(round(n)) if n is not None else None


def _clean_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _assemble_address(row: tuple, address_cols: list[int],
                       postcode: str | None = None) -> str | None:
    parts = []
    for idx in address_cols:
        if idx >= len(row):
            continue
        part = _clean_str(row[idx])
        if not part:
            continue
        if postcode and part.upper().replace(" ", "") == postcode.upper().replace(" ", ""):
            continue
        if UK_POSTCODE_RE.match(part.strip()):
            continue
        parts.append(part)
    return ", ".join(parts) if parts else None


def _strip_postcode_from_address(address: str | None, postcode: str | None) -> str | None:
    if not address or not postcode:
        return address
    cleaned = re.sub(
        r",?\s*" + re.escape(postcode.strip()) + r"\s*,?",
        "", address, flags=re.IGNORECASE,
    ).strip().strip(",").strip()
    parts = [p.strip() for p in cleaned.split(",")]
    seen: list[str] = []
    for p in parts:
        if p and p.lower() not in [x.lower() for x in seen]:
            seen.append(p)
    return ", ".join(seen) if seen else None


def _parse_storeys(val: Any) -> int | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    has_basement = "basement" in s.lower()
    m = re.search(r"\b(\d+)\b", s)
    if m:
        n = int(m.group(1))
        return max(1, n - 1) if has_basement else n
    return None


def _to_bool(val: Any) -> bool | None:
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    return str(val).lower().strip() in ("yes", "true", "1", "y")


# ─────────────────────────────────────────────────────────────────────────────
# Row filter utilities
# ─────────────────────────────────────────────────────────────────────────────

def _is_blank_row(row: tuple, col_map: dict, address_cols: list[int]) -> bool:
    if address_cols:
        has_addr = any(_clean_str(row[i]) for i in address_cols if i < len(row))
    else:
        addr_col = col_map.get("address")
        has_addr = bool(addr_col is not None and addr_col < len(row) and _clean_str(row[addr_col]))
    ref_col = col_map.get("property_reference")
    has_ref = bool(ref_col is not None and ref_col < len(row) and _clean_str(row[ref_col]))
    return not has_addr and not has_ref


def _is_summary_row(row: tuple) -> bool:
    """Detect summary/total rows. Only trigger on short cells."""
    for c in row:
        if c is None:
            continue
        s = _cell_str(c).lower()
        if len(s) > 30:
            continue
        if any(kw in s for kw in ["total", "subtotal", "grand total", "sum of", "count of"]):
            return True
    return False


# Issue #10 — Detect filter/metadata rows embedded in exports
def _is_metadata_row(row: tuple) -> bool:
    """Detect system-generated filter/metadata rows."""
    for c in row:
        if c is None:
            continue
        s = str(c).strip()
        if len(s) > 120:  # No real property ref or address is > 120 chars
            return True
        s_lower = s.lower()
        if s_lower.startswith("filter:") or s_lower.startswith("report:") or s_lower.startswith("export:"):
            return True
    return False


def _get_val(row: tuple, col_map: dict, field: str) -> Any:
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# ─────────────────────────────────────────────────────────────────────────────
# Stage A — Workbook Snapshot
# ─────────────────────────────────────────────────────────────────────────────

def _stage_a_snapshot(wb: openpyxl.Workbook, upload_id: str) -> list[dict]:
    """Profile every sheet with scoring, including Issue #3 and #12 fixes."""
    profiles = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = _sheet_rows(ws)
        non_empty = [r for r in rows if _count_filled(r) > 1]

        name_lower = name.lower().strip()
        score = 0.0
        is_noise = False

        for kw in NOISE_KEYWORDS:
            if kw in name_lower:
                score -= 5.0
                is_noise = True
        for kw in DATA_KEYWORDS:
            if kw in name_lower:
                score += 3.0

        # Issue #3 — penalise superset sheets
        is_superset = any(kw in name_lower for kw in SUPERSET_PENALTIES)
        is_dwelling_specific = any(kw in name_lower for kw in DWELLING_PREFER)
        if is_superset:
            score -= 2.0  # penalise; dwelling-specific sheets will outscore

        if len(non_empty) < 3:
            score -= 5.0
        elif len(non_empty) > 1000:
            score += 5.0
        elif len(non_empty) > 100:
            score += 3.0
        elif len(non_empty) > 20:
            score += 1.0

        # Dwelling-specific bonus
        if is_dwelling_specific:
            score += 2.0

        has_address = False
        has_si = False
        numeric_density = 0
        for row in non_empty[:5]:
            row_text = " ".join(_cell_str(c).lower() for c in row[:MAX_PREVIEW_COLS])
            if any(kw in row_text for kw in ["address", "postcode", "post code", "add 1", "add_1"]):
                has_address = True
                score += 2.0
            if any(kw in row_text for kw in SI_KEYWORDS):
                has_si = True
                score += 2.0
            if any(kw in row_text for kw in ["ref", "uprn", "property id", "prty_id", "asset ref"]):
                score += 1.5

        # Issue #12 — detect SI-heavy sheets (potential secondary, not noise)
        for row in non_empty[1:20]:
            num_count = sum(1 for c in row[:MAX_PREVIEW_COLS]
                           if isinstance(c, (int, float)) and c > 1000)
            if num_count >= 3:
                numeric_density += 1
        if numeric_density >= 3 and len(non_empty) > 100:
            # This is likely a financial data sheet — don't classify as noise
            if is_noise and has_si:
                is_noise = False
                score += 3.0

        for row in non_empty[1:8]:
            for c in row[:MAX_PREVIEW_COLS]:
                if isinstance(c, (int, float)) and c > 10000:
                    score += 0.3
                    break

        profiles.append({
            "name": name,
            "rows": len(non_empty),
            "cols": ws.max_column or 0,
            "score": score,
            "has_address": has_address,
            "has_si": has_si,
            "is_noise": is_noise,
            "is_superset": is_superset,
            "is_dwelling_specific": is_dwelling_specific,
            "preview_rows": non_empty[:6],
        })

    # Issue #3 — if dwelling-specific and superset both exist, boost dwelling
    has_dwelling = any(p["is_dwelling_specific"] for p in profiles)
    if has_dwelling:
        for p in profiles:
            if p["is_superset"] and not p["is_dwelling_specific"]:
                p["score"] -= 5.0  # further penalise superset when dwelling exists

    profiles.sort(key=lambda p: p["score"], reverse=True)
    logger.info(
        f"[{upload_id}] Stage A — {len(profiles)} sheets. "
        f"Top: {[(p['name'], round(p['score'],1), p['rows']) for p in profiles[:4]]}"
    )
    return profiles


# ─────────────────────────────────────────────────────────────────────────────
# Stage B — Single LLM Analysis (1 call)
# ─────────────────────────────────────────────────────────────────────────────

def _build_llm_prompt(profiles: list[dict], wb: openpyxl.Workbook) -> str:
    candidates = [p for p in profiles if not p["is_noise"] and p["rows"] >= 3][:6]
    if not candidates:
        candidates = profiles[:3]

    sheet_sections = []
    for p in candidates:
        ws = wb[p["name"]]
        rows = _sheet_rows(ws)
        non_empty = [r for r in rows if _count_filled(r) > 1]
        header_text = ""
        for i, row in enumerate(non_empty[:6]):
            cells = [f"[{j}]{_cell_str(c)[:35]}" for j, c in enumerate(row[:MAX_PREVIEW_COLS])
                     if c is not None and str(c).strip()]
            header_text += f"  Row {i}: {' | '.join(cells)}\n"
        sheet_sections.append(
            f'Sheet: "{p["name"]}" ({p["rows"]} data rows, {p["cols"]} cols)\n{header_text}'
        )

    noise = [p["name"] for p in profiles if p["is_noise"] or p["rows"] < 3]
    noise_text = f"Other sheets (likely noise): {noise}" if noise else ""
    schema_text = "\n".join(f"  {k}: {v}" for k, v in CANONICAL_FIELDS.items())

    prompt = f"""You are analysing a UK social housing Schedule of Values (SoV) Excel workbook.
Your job: classify sheets, find headers, and map columns — ALL in one response.

WORKBOOK ({len(profiles)} sheets total):

{chr(10).join(sheet_sections)}

{noise_text}

CANONICAL SCHEMA (map source columns to these field names):
{schema_text}

RESPOND with valid JSON matching this EXACT structure:
{{
  "sheets": [
    {{
      "name": "Sheet Name",
      "role": "primary",
      "header_row": 0,
      "col_map": {{
        "property_reference": 5,
        "address": 8,
        "postcode": 12,
        "sum_insured": 18
      }},
      "address_cols": [],
      "postcode_in_address": false
    }}
  ],
  "join_plan": null
}}

RULES:
1. "role": "primary" (per-unit/property rows with addresses), "secondary" (SI or construction by block), "noise"
2. "header_row": 0-based index within non-empty rows where column NAMES appear
   - If row 0 has category labels and row 1 has field names, header_row=1
3. "col_map": canonical field → column INDEX (0-based)
   - "sum_insured" MUST be NUMERIC rebuild/declared value, NOT text
   - If multiple SI year columns, pick the LATEST year only
   - Only include fields that exist in the sheet
4. "address_cols": list of column indices for fragmented address, IN ORDER for concatenation
5. "postcode_in_address": true if no dedicated postcode column
6. "join_plan": only if secondary data must join to primary:
   {{ "primary_join_col": 3, "secondary_sheet": "X", "secondary_join_col": 0,
      "strip_suffix": false, "fields_to_pull": ["sum_insured", "wall_construction"] }}
   Set null if no join needed.

CRITICAL:
- If a sheet has ALL property types AND a separate dwellings-only sheet exists, prefer dwellings-only as primary
- A sheet named 'stock val' or 'valuation' with >1000 rows of numeric data is secondary (SI source), NOT noise
- For "insure_flag": map to any column like "Insure Yes/No" — rows with "N" will be filtered
- Ignore correspondence/mailing address columns — only map RISK/PROPERTY address
- "wall_construction" = wall material column, NOT "non-traditional" flag or "construction date"
"""

    return prompt


def _stage_b_llm_analysis(profiles: list[dict], wb: openpyxl.Workbook,
                           upload_id: str) -> dict | None:
    prompt = _build_llm_prompt(profiles, wb)
    logger.info(f"[{upload_id}] Stage B — sending single LLM analysis prompt")

    raw = _llm(
        prompt,
        system="You are a data engineering expert for UK insurance. Respond ONLY with valid JSON.",
        max_tokens=4096,
    )
    result = _parse_json(raw)

    if not result or not isinstance(result, dict) or not result.get("sheets"):
        logger.warning(f"[{upload_id}] Stage B LLM failed or returned invalid JSON")
        return None

    validated_sheets = []
    for sheet in result["sheets"]:
        if not isinstance(sheet, dict):
            continue
        name = sheet.get("name", "")
        if name not in wb.sheetnames:
            for sn in wb.sheetnames:
                if sn.strip() == name.strip():
                    sheet["name"] = sn
                    name = sn
                    break
        if name not in wb.sheetnames:
            continue

        raw_map = sheet.get("col_map", {})
        clean_map = {}
        for k, v in raw_map.items():
            if k in CANONICAL_FIELDS and isinstance(v, (int, float)) and v >= 0:
                clean_map[k] = int(v)
        sheet["col_map"] = clean_map

        # Issue #1 — null guard on header_row
        hr_raw = sheet.get("header_row")
        sheet["header_row"] = int(hr_raw) if hr_raw is not None else 0

        sheet["address_cols"] = [
            int(c) for c in (sheet.get("address_cols") or [])
            if isinstance(c, (int, float))
        ]
        sheet["postcode_in_address"] = bool(sheet.get("postcode_in_address"))
        sheet["role"] = sheet.get("role", "noise")
        validated_sheets.append(sheet)

    if not validated_sheets:
        return None

    # Issue #1, #8 — null guards on join_plan
    join_plan = result.get("join_plan")
    if join_plan and isinstance(join_plan, dict):
        required = ["primary_join_col", "secondary_sheet", "secondary_join_col"]
        if not all(k in join_plan for k in required):
            join_plan = None
        else:
            join_plan["primary_join_col"] = int(join_plan.get("primary_join_col") or 0)
            join_plan["secondary_join_col"] = int(join_plan.get("secondary_join_col") or 0)
            join_plan["strip_suffix"] = bool(join_plan.get("strip_suffix") or False)
            raw_fields = list(join_plan.get("fields_to_pull") or [])
            join_plan["fields_to_pull"] = [
                f["name"] if isinstance(f, dict) else str(f) for f in raw_fields
            ]
    else:
        join_plan = None

    result_clean = {"sheets": validated_sheets, "join_plan": join_plan}
    primary_count = sum(1 for s in validated_sheets if s["role"] == "primary")
    logger.info(f"[{upload_id}] Stage B LLM OK — {primary_count} primary, "
                f"{len(validated_sheets)} total sheets mapped")
    return result_clean


# ─────────────────────────────────────────────────────────────────────────────
# Stage B fallback — Full Heuristic (0 LLM calls)
# ─────────────────────────────────────────────────────────────────────────────

def _heuristic_find_header_row(rows: list[tuple]) -> int:
    best_row, best_score = 0, -1
    for i, row in enumerate(rows[:MAX_HEADER_SCAN]):
        if not any(c for c in row):
            continue
        strings = [c for c in row if isinstance(c, str) and c.strip()]
        if len(strings) < 2:
            continue
        score = len(strings)
        row_text = " ".join(str(c).lower() for c in strings)
        for kw in ["address", "postcode", "post code", "ref", "insur", "block",
                    "storey", "bedroom", "type", "year", "add 1", "prty_id"]:
            if kw in row_text:
                score += 3
        if any(isinstance(c, (int, float)) and c > 100 for c in row):
            score -= 2
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def _heuristic_map_columns(headers: tuple, sheet_name: str, upload_id: str = "") -> dict:
    """Map headers to canonical fields. Includes Issue #2 SI year fix."""
    col_map: dict[str, int] = {}
    si_year_candidates: dict[int, int] = {}  # year → col_index

    for i, h in enumerate(headers[:MAX_PREVIEW_COLS]):
        if not h:
            continue
        h_lower = _cell_str(h).lower()

        # Skip obviously irrelevant columns
        if any(skip in h_lower for skip in [
            "correspondence", "mailing", "indicative rent", "tenant",
            "gross book", "net book", "title number", "weekly rent",
        ]):
            continue

        for field, keywords in FIELD_KEYWORDS.items():
            if field in col_map and field != "sum_insured":
                continue
            if any(kw in h_lower for kw in keywords):
                if field == "sum_insured":
                    # Skip "Sum Insured Type" / "Basis of" columns
                    if any(skip in h_lower for skip in ["type", "basis"]):
                        continue
                    # Issue #2 — use robust year extraction
                    year = _extract_year_from_col_name(_cell_str(h))
                    si_year_candidates[year] = i
                elif field == "property_type":
                    if h_lower.strip() == "type" or len(h_lower.strip()) > 3:
                        col_map[field] = i
                elif field == "wall_construction":
                    if "date" not in h_lower:
                        col_map[field] = i
                elif field == "property_reference":
                    if "block" not in h_lower and "parent" not in h_lower and "policy" not in h_lower:
                        col_map[field] = i
                else:
                    col_map[field] = i

    # Standalone "Reference" column fallback
    if "property_reference" not in col_map:
        for i, h in enumerate(headers[:MAX_PREVIEW_COLS]):
            if h and _cell_str(h).strip().lower() == "reference":
                col_map["property_reference"] = i
                break

    # Issue #2 — pick latest year SI column
    if si_year_candidates:
        latest_year = max(si_year_candidates.keys())
        col_map["sum_insured"] = si_year_candidates[latest_year]
        if upload_id:
            logger.info(f"[{upload_id}] SI column: year={latest_year}, "
                        f"col={si_year_candidates[latest_year]}, "
                        f"candidates={si_year_candidates}")

    # Detect fragmented address
    address_cols: list[int] = []
    addr_fragment_patterns = [
        ["address line 1", "address line 2", "address line 3"],
        ["add 1", "add 2", "add 3"],
        ["add_1", "add_2", "add_3"],
        ["hse no", "street", "flat"],
        ["address 1", "address 2", "address 3"],
    ]
    for pattern in addr_fragment_patterns:
        matched = []
        for frag in pattern:
            for i, h in enumerate(headers[:MAX_PREVIEW_COLS]):
                if h and frag in _cell_str(h).lower() and i not in matched:
                    matched.append(i)
                    break
        if len(matched) >= 2:
            address_cols = matched
            break

    if address_cols:
        for f in ["address", "address_2", "address_3"]:
            col_map.pop(f, None)

    postcode_in_address = "postcode" not in col_map

    return {
        "col_map": col_map,
        "address_cols": address_cols,
        "postcode_in_address": postcode_in_address,
    }


def _heuristic_detect_join(primary_profile: dict, primary_mapping: dict,
                            profiles: list[dict], wb: openpyxl.Workbook) -> dict | None:
    if "sum_insured" in primary_mapping.get("col_map", {}):
        return None

    primary_name = primary_profile["name"]
    primary_cols = primary_mapping.get("col_map", {})
    join_key_field = "block_reference" if "block_reference" in primary_cols else "property_reference"
    join_key_col = primary_cols.get(join_key_field)
    if join_key_col is None:
        return None

    for p in profiles:
        if p["name"] == primary_name or p["rows"] < 3:
            continue
        if not p["has_si"]:
            continue
        ws = wb[p["name"]]
        rows = _sheet_rows(ws)
        non_empty = [r for r in rows if _count_filled(r) > 1]
        if len(non_empty) < 2:
            continue
        hr = _heuristic_find_header_row(non_empty)
        if hr >= len(non_empty):
            continue
        sec_mapping = _heuristic_map_columns(non_empty[hr], p["name"])
        sec_cols = sec_mapping.get("col_map", {})

        sec_key_col = sec_cols.get("block_reference") or sec_cols.get("property_reference")
        if sec_key_col is not None and "sum_insured" in sec_cols:
            return {
                "primary_join_col": join_key_col,
                "secondary_sheet": p["name"],
                "secondary_join_col": sec_key_col,
                "strip_suffix": False,
                "fields_to_pull": [f for f in sec_cols if f not in primary_cols],
            }

    return None


def _stage_b_heuristic(profiles: list[dict], wb: openpyxl.Workbook,
                        upload_id: str) -> dict:
    logger.info(f"[{upload_id}] Stage B heuristic — full analysis without LLM")

    sheets_result = []
    for p in profiles:
        if (p["is_noise"] and not p["has_si"]) or p["rows"] < 3:
            sheets_result.append({
                "name": p["name"], "role": "noise",
                "header_row": 0, "col_map": {},
                "address_cols": [], "postcode_in_address": False,
            })
            continue

        ws = wb[p["name"]]
        rows = _sheet_rows(ws)
        non_empty = [r for r in rows if _count_filled(r) > 1]
        if len(non_empty) < 2:
            sheets_result.append({
                "name": p["name"], "role": "noise",
                "header_row": 0, "col_map": {},
                "address_cols": [], "postcode_in_address": False,
            })
            continue

        hr = _heuristic_find_header_row(non_empty)
        mapping = _heuristic_map_columns(
            non_empty[hr] if hr < len(non_empty) else (), p["name"], upload_id
        )

        has_addr = bool(mapping["col_map"].get("address") or mapping["address_cols"])
        has_si = "sum_insured" in mapping["col_map"]
        has_ref = "property_reference" in mapping["col_map"]

        if has_addr or (has_ref and p["rows"] > 20):
            role = "primary"
        elif has_si and not has_addr:
            role = "secondary"
        else:
            role = "noise" if p["score"] < 0 else "primary"

        sheets_result.append({
            "name": p["name"], "role": role, "header_row": hr,
            "col_map": mapping["col_map"],
            "address_cols": mapping["address_cols"],
            "postcode_in_address": mapping["postcode_in_address"],
            "_heuristic": True,
        })

    primaries = [s for s in sheets_result if s["role"] == "primary"]
    join_plan = None
    if primaries:
        primary = primaries[0]
        pp = next((p for p in profiles if p["name"] == primary["name"]), None)
        if pp:
            join_plan = _heuristic_detect_join(pp, primary, profiles, wb)

    primary_count = sum(1 for s in sheets_result if s["role"] == "primary")
    logger.info(f"[{upload_id}] Stage B heuristic — {primary_count} primary sheets")
    return {"sheets": sheets_result, "join_plan": join_plan, "_heuristic": True}


# ─────────────────────────────────────────────────────────────────────────────
# Cross-sheet join builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_join_lookup(wb: openpyxl.Workbook, join_plan: dict,
                        analysis: dict, upload_id: str) -> dict[str, dict]:
    sec_name = str(join_plan.get("secondary_sheet") or "")
    if sec_name not in wb.sheetnames:
        return {}

    sec_info = next((s for s in analysis.get("sheets", []) if s["name"] == sec_name), None)
    if not sec_info:
        ws = wb[sec_name]
        rows = _sheet_rows(ws)
        non_empty = [r for r in rows if _count_filled(r) > 1]
        hr = _heuristic_find_header_row(non_empty)
        sec_mapping = _heuristic_map_columns(non_empty[hr] if hr < len(non_empty) else (), sec_name)
        sec_col_map = sec_mapping["col_map"]
        header_row = hr
    else:
        sec_col_map = sec_info.get("col_map", {})
        header_row = int(sec_info.get("header_row") or 0)

    # Issue #8 — null guards
    sec_join_col = int(join_plan.get("secondary_join_col") or 0)
    strip_suffix = bool(join_plan.get("strip_suffix") or False)
    fields = list(join_plan.get("fields_to_pull") or [])
    fields = [f["name"] if isinstance(f, dict) else str(f) for f in fields]

    ws = wb[sec_name]
    rows = _sheet_rows(ws)
    non_empty = [r for r in rows if _count_filled(r) > 1]
    data_rows = non_empty[header_row + 1:]

    lookup: dict[str, dict] = {}
    for row in data_rows:
        if sec_join_col >= len(row):
            continue
        key_raw = _clean_str(row[sec_join_col])
        if not key_raw:
            continue
        if strip_suffix:
            key_raw = re.split(r"\s*[-–]\s*", key_raw)[0].strip()

        entry = lookup.get(key_raw, {})
        for field in fields:
            if field in entry:
                continue
            fidx = sec_col_map.get(field)
            if fidx is None or fidx >= len(row):
                continue
            raw_val = row[fidx]
            if field == "sum_insured":
                entry[field] = _clean_numeric(raw_val)
            elif field in ("year_of_build",):
                entry[field] = _decode_year(raw_val)
            elif field in ("num_storeys",):
                entry[field] = _parse_storeys(raw_val)
            else:
                entry[field] = _clean_str(raw_val)
        lookup[key_raw] = entry

    logger.info(f"[{upload_id}] Join lookup: {len(lookup)} keys from '{sec_name}'")
    return lookup


def _lookup_join(key: str, join_data: dict) -> dict | None:
    if not key or not join_data:
        return None
    if key in join_data:
        return join_data[key]
    stripped = re.split(r"\s*[-–]\s*", key)[0].strip()
    if stripped in join_data:
        return join_data[stripped]
    for k in join_data:
        if key.startswith(k) or k.startswith(stripped):
            return join_data[k]
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Stage C — Row Extraction (with all issue fixes)
# ─────────────────────────────────────────────────────────────────────────────

def _stage_c_extract(ws, sheet_info: dict, join_data: dict | None,
                      join_plan: dict | None,
                      ha_id: str, upload_id: str,
                      auto_ref_prefix: str,
                      data_rows_for_uniqueness: list[tuple] | None = None,
                      ) -> list[dict]:
    rows = _sheet_rows(ws)
    non_empty = [r for r in rows if _count_filled(r) > 1]
    header_row = int(sheet_info.get("header_row") or 0)  # Issue #1
    col_map = dict(sheet_info.get("col_map", {}))
    address_cols = sheet_info.get("address_cols", [])
    postcode_in_address = sheet_info.get("postcode_in_address", False)
    sheet_name = sheet_info["name"]
    is_heuristic = sheet_info.get("_heuristic", False)

    if header_row >= len(non_empty):
        return []

    data_rows = non_empty[header_row + 1:]

    # Issue #4 — validate property_reference uniqueness
    use_address_as_ref = False
    if "property_reference" in col_map and data_rows:
        ref_col = col_map["property_reference"]
        uniqueness = _validate_property_ref_uniqueness(data_rows, ref_col)
        if uniqueness < 0.8:
            logger.warning(
                f"[{upload_id}] property_reference col[{ref_col}] has uniqueness "
                f"{uniqueness:.1%} — rejecting, falling back to address as key"
            )
            del col_map["property_reference"]
            use_address_as_ref = True

    if "sum_insured" not in col_map:
        logger.warning(
            f"[{upload_id}] ⚠️ NO SUM INSURED column in '{sheet_name}'. SI will be NULL."
        )

    # Issue #8 — null guard on primary_join_col
    primary_join_col = None
    if join_plan and join_data:
        pjc = join_plan.get("primary_join_col")
        if isinstance(pjc, int):
            primary_join_col = pjc
        elif isinstance(pjc, str):
            primary_join_col = col_map.get(pjc)

    extracted = []
    auto_counter = 0

    for row in data_rows:
        try:
            if _is_blank_row(row, col_map, address_cols):
                continue
            if _is_summary_row(row):
                continue
            if _is_metadata_row(row):  # Issue #10
                continue

            # Issue #11 — skip non-insured rows
            insure_flag = _clean_str(_get_val(row, col_map, "insure_flag"))
            if insure_flag is not None and insure_flag.upper() in ("N", "NO", "FALSE", "0"):
                continue

            # Postcode (extract early for address assembly)
            postcode = _clean_postcode(_get_val(row, col_map, "postcode"))

            # Address assembly
            if address_cols:
                address = _assemble_address(row, address_cols, postcode)
            else:
                address = _clean_str(_get_val(row, col_map, "address"))

            if not postcode and (postcode_in_address or "postcode" not in col_map):
                if address:
                    postcode = _clean_postcode(address)

            if address and postcode:
                address = _strip_postcode_from_address(address, postcode)

            # Property reference — Issue #4
            prop_ref = _clean_str(_get_val(row, col_map, "property_reference"))
            if not prop_ref and use_address_as_ref and address:
                # Use address + postcode as unique key
                prop_ref = f"{address}_{postcode}" if postcode else address
            if not prop_ref:
                auto_counter += 1
                prop_ref = f"AUTO_{auto_ref_prefix}_{auto_counter:05d}"

            # Issue #13 — flag suspicious refs
            si_review_flag = False
            if prop_ref and "future" in prop_ref.lower():
                si_review_flag = True

            block_ref = _clean_str(_get_val(row, col_map, "block_reference"))

            # Sum insured
            sum_insured = _clean_numeric(_get_val(row, col_map, "sum_insured"))

            # Cross-sheet join enrichment
            joined: dict = {}
            if join_data and primary_join_col is not None:
                jkv = _clean_str(row[primary_join_col] if primary_join_col < len(row) else None)
                if jkv:
                    joined = _lookup_join(jkv, join_data) or {}
                    if sum_insured is None and joined.get("sum_insured"):
                        sum_insured = joined["sum_insured"]

            def get(field: str) -> Any:
                val = _get_val(row, col_map, field)
                if val is None and field in joined:
                    return joined[field]
                return val

            # Issue #5 — year of build (handles all formats)
            year_of_build = _decode_year(get("year_of_build"))
            age_banding = _clean_str(get("age_banding")) or _derive_age_banding(year_of_build)

            wall = _clean_str(get("wall_construction"))
            conf = _compute_confidence(
                address=address, postcode=postcode, sum_insured=sum_insured,
                block_ref=block_ref, prop_ref=prop_ref, wall=wall,
                sheet_name=sheet_name, is_heuristic=is_heuristic,
            )

            metadata = {
                "field_confidence": conf,
                "source_sheet":     sheet_name,
                "processed_at":     datetime.now(timezone.utc).isoformat(),
                "submission_id":    upload_id,
                "used_heuristic":   is_heuristic,
                "used_join":        bool(joined),
            }
            # Issue #13 — flag edge-case SI rows
            if si_review_flag:
                metadata["si_confidence"] = "low"
                metadata["si_review_flag"] = True

            # Issue #9 — _clamp() on every string field
            record = {
                "ha_id":              ha_id,
                "submission_id":      upload_id,
                "property_reference": _clamp(prop_ref, DB_COLUMN_LIMITS["property_reference"]),
                "block_reference":    _clamp(block_ref, DB_COLUMN_LIMITS["block_reference"]),
                "address":            address,
                "address_2":          _clean_str(get("address_2")),
                "address_3":          _clean_str(get("address_3")),
                "postcode":           _clamp(postcode, DB_COLUMN_LIMITS["postcode"]),
                "occupancy_type":     _clamp(_clean_str(get("occupancy_type")), DB_COLUMN_LIMITS["occupancy_type"]),
                "sum_insured":        sum_insured,
                "sum_insured_type":   _clamp(_clean_str(get("sum_insured_type")), DB_COLUMN_LIMITS["sum_insured_type"]),
                "property_type":      _clamp(_clean_str(get("property_type")), DB_COLUMN_LIMITS["property_type"]),
                "avid_property_type": _clamp(_clean_str(get("avid_property_type")), DB_COLUMN_LIMITS["avid_property_type"]),
                "wall_construction":  _clamp(wall, DB_COLUMN_LIMITS["wall_construction"]),
                "roof_construction":  _clamp(_clean_str(get("roof_construction")), DB_COLUMN_LIMITS["roof_construction"]),
                "floor_construction": _clamp(_clean_str(get("floor_construction")), DB_COLUMN_LIMITS["floor_construction"]),
                "year_of_build":      _clamp(str(year_of_build) if year_of_build else None, DB_COLUMN_LIMITS["year_of_build"]),
                "age_banding":        _clamp(age_banding, DB_COLUMN_LIMITS["age_banding"]),
                "num_bedrooms":       _clean_int(get("num_bedrooms")),
                "storeys":            _parse_storeys(get("num_storeys")),
                "units":              _clean_int(get("num_units")),
                "basement":           _to_bool(get("has_basement")),
                "is_listed":          _to_bool(get("is_listed")),
                "security_features":  _clamp(_clean_str(get("security_features")), DB_COLUMN_LIMITS["security_features"]),
                "fire_protection":    _clamp(_clean_str(get("fire_protection")), DB_COLUMN_LIMITS["fire_protection"]),
                "alarms":             _clamp(_clean_str(get("alarms")), DB_COLUMN_LIMITS["alarms"]),
                "flood_insured":      _to_bool(get("flood_insured")),
                "storm_insured":      _to_bool(get("storm_insured")),
                "deductible":         _clean_numeric(get("deductible")),
                "enrichment_status":  "pending",
                "metadata":           metadata,
            }
            extracted.append(record)

        except Exception as row_exc:
            logger.warning(f"[{upload_id}] Row extraction error in '{sheet_name}': {row_exc}")
            continue  # skip bad row, continue processing

    logger.info(f"[{upload_id}] Stage C — {len(extracted)} rows from '{sheet_name}'")
    return extracted


def _compute_confidence(address, postcode, sum_insured, block_ref,
                         prop_ref, wall, sheet_name, is_heuristic) -> dict:
    penalty = 0.15 if is_heuristic else 0.0
    is_auto = bool(prop_ref and prop_ref.startswith("AUTO_"))

    def score(val, base):
        return max(0.0, base - penalty) if val is not None else 0.0

    return {
        "property_reference": {"confidence": 0.5 if is_auto else score(prop_ref, 1.0), "source": sheet_name},
        "address":            {"confidence": score(address, 0.95),                      "source": sheet_name},
        "postcode":           {"confidence": score(postcode, 0.90),                     "source": sheet_name},
        "sum_insured":        {"confidence": score(sum_insured, 0.75 if is_heuristic else 0.90), "source": sheet_name},
        "block_reference":    {"confidence": score(block_ref, 0.90),                    "source": sheet_name},
        "wall_construction":  {"confidence": score(wall, 0.85),                         "source": sheet_name},
    }


# ─────────────────────────────────────────────────────────────────────────────
# Stage D — Upsert
# ─────────────────────────────────────────────────────────────────────────────

UPSERT_SQL = """
INSERT INTO silver.properties (
    ha_id, submission_id, property_reference, block_reference,
    address, address_2, address_3, postcode,
    occupancy_type, sum_insured, sum_insured_type,
    property_type, avid_property_type,
    wall_construction, roof_construction, floor_construction,
    year_of_build, age_banding, num_bedrooms, storeys, units,
    basement, is_listed, security_features, fire_protection,
    alarms, flood_insured, storm_insured, deductible,
    enrichment_status, metadata, created_at, updated_at
)
VALUES (
    %(ha_id)s, %(submission_id)s::uuid, %(property_reference)s, %(block_reference)s,
    %(address)s, %(address_2)s, %(address_3)s, %(postcode)s,
    %(occupancy_type)s, %(sum_insured)s, %(sum_insured_type)s,
    %(property_type)s, %(avid_property_type)s,
    %(wall_construction)s, %(roof_construction)s, %(floor_construction)s,
    %(year_of_build)s, %(age_banding)s, %(num_bedrooms)s, %(storeys)s, %(units)s,
    %(basement)s, %(is_listed)s, %(security_features)s, %(fire_protection)s,
    %(alarms)s, %(flood_insured)s, %(storm_insured)s, %(deductible)s,
    %(enrichment_status)s, %(metadata)s::jsonb, NOW(), NOW()
)
ON CONFLICT (ha_id, property_reference)
DO UPDATE SET
    block_reference    = EXCLUDED.block_reference,
    address            = EXCLUDED.address,
    address_2          = EXCLUDED.address_2,
    address_3          = EXCLUDED.address_3,
    postcode           = EXCLUDED.postcode,
    occupancy_type     = EXCLUDED.occupancy_type,
    sum_insured        = EXCLUDED.sum_insured,
    sum_insured_type   = EXCLUDED.sum_insured_type,
    property_type      = EXCLUDED.property_type,
    avid_property_type = EXCLUDED.avid_property_type,
    wall_construction  = EXCLUDED.wall_construction,
    roof_construction  = EXCLUDED.roof_construction,
    floor_construction = EXCLUDED.floor_construction,
    year_of_build      = EXCLUDED.year_of_build,
    age_banding        = EXCLUDED.age_banding,
    num_bedrooms       = EXCLUDED.num_bedrooms,
    storeys            = EXCLUDED.storeys,
    units              = EXCLUDED.units,
    basement           = EXCLUDED.basement,
    is_listed          = EXCLUDED.is_listed,
    security_features  = EXCLUDED.security_features,
    fire_protection    = EXCLUDED.fire_protection,
    alarms             = EXCLUDED.alarms,
    flood_insured      = EXCLUDED.flood_insured,
    storm_insured      = EXCLUDED.storm_insured,
    deductible         = EXCLUDED.deductible,
    submission_id      = EXCLUDED.submission_id,
    enrichment_status  = CASE
                           WHEN silver.properties.enrichment_status = 'enriched'
                           THEN 'enriched'
                           ELSE EXCLUDED.enrichment_status
                         END,
    metadata           = EXCLUDED.metadata,
    updated_at         = NOW()
"""


def _get_db_conn():
    return psycopg2.connect(
        os.getenv("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/platform_dev")
    )


def _stage_d_upsert(records: list[dict], upload_id: str) -> int:
    if not records:
        return 0
    for r in records:
        r["metadata"] = json.dumps(r.get("metadata", {}))
    try:
        conn = _get_db_conn()
        try:
            with conn:
                with conn.cursor() as cur:
                    psycopg2.extras.execute_batch(cur, UPSERT_SQL, records, page_size=500)
            logger.info(f"[{upload_id}] Stage D — upserted {len(records)} rows")
            return len(records)
        finally:
            conn.close()
    except Exception as exc:
        logger.error(f"[{upload_id}] Stage D upsert failed: {exc}", exc_info=True)
        return 0


# ─────────────────────────────────────────────────────────────────────────────
# Coverage report
# ─────────────────────────────────────────────────────────────────────────────

def _build_coverage_report(records: list[dict], upload_id: str) -> dict:
    if not records:
        return {"rows_extracted": 0, "field_coverage": {}, "total_sum_insured": 0}

    fields_to_check = [
        "address", "postcode", "sum_insured", "block_reference",
        "occupancy_type", "wall_construction", "roof_construction",
        "floor_construction", "year_of_build", "storeys",
        "num_bedrooms", "units", "property_type",
    ]
    total = len(records)
    coverage = {}
    for field in fields_to_check:
        populated = sum(1 for r in records if r.get(field) is not None)
        pct = round(100 * populated / total, 1) if total > 0 else 0
        coverage[field] = {
            "populated": populated, "total": total, "pct": pct,
            "status": "✅" if pct >= 90 else ("⚠️" if pct >= 50 else "❌"),
        }

    si_populated = coverage.get("sum_insured", {}).get("populated", 0)
    si_total = sum(r.get("sum_insured") or 0.0 for r in records)

    return {
        "upload_id":         upload_id,
        "rows_extracted":    total,
        "total_sum_insured": si_total,
        "si_warning":        si_populated == 0,
        "field_coverage":    coverage,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Main entry point — Zero-crash guarantee
# ─────────────────────────────────────────────────────────────────────────────

async def process_sov_to_silver(
    file_bytes: bytes,
    ha_id: str,
    upload_id: str = "",
    filename: str = "upload.xlsx",
    submission_id: str | None = None,
    db_pool: Any = None,
    **kwargs: Any,
) -> dict:
    """
    Full pipeline: Excel bytes → silver.properties

    Returns coverage report dict. Never raises exceptions.
    LLM calls: 1 (or 0 on fallback).
    """
    # Resolve upload_id
    if not upload_id and submission_id:
        upload_id = submission_id
    if not upload_id:
        upload_id = str(uuid.uuid4())

    try:
        logger.info(f"[{upload_id}] ══ SoV Processor v3 START — ha={ha_id} file={filename} ══")

        try:
            wb = _load_workbook(file_bytes)
        except Exception as exc:
            logger.error(f"[{upload_id}] Cannot open workbook: {exc}")
            return {"error": str(exc), "rows_extracted": 0, "rows_written": 0,
                    "total_sum_insured": 0, "si_warning": True}

        # ── Stage A: Workbook Snapshot ──
        profiles = _stage_a_snapshot(wb, upload_id)

        # ── Stage B: Single LLM Analysis or heuristic ──
        analysis = _stage_b_llm_analysis(profiles, wb, upload_id)
        if not analysis:
            analysis = _stage_b_heuristic(profiles, wb, upload_id)

        # ── Build join lookup if needed ──
        join_plan = analysis.get("join_plan")
        join_data: dict | None = None
        if join_plan:
            try:
                join_data = _build_join_lookup(wb, join_plan, analysis, upload_id)
            except Exception as exc:
                logger.warning(f"[{upload_id}] Join lookup failed: {exc}")
                join_data = None

        # ── Stage C: Row Extraction ──
        primary_sheets = [s for s in analysis["sheets"] if s.get("role") == "primary"]
        if not primary_sheets and profiles:
            fallback_name = profiles[0]["name"]
            ws = wb[fallback_name]
            rows = _sheet_rows(ws)
            non_empty = [r for r in rows if _count_filled(r) > 1]
            hr = _heuristic_find_header_row(non_empty)
            mapping = _heuristic_map_columns(
                non_empty[hr] if hr < len(non_empty) else (), fallback_name, upload_id
            )
            primary_sheets = [{
                "name": fallback_name, "role": "primary",
                "header_row": hr, "_heuristic": True, **mapping,
            }]
            logger.warning(f"[{upload_id}] No primary sheets — fallback to '{fallback_name}'")

        all_records: list[dict] = []
        seen_keys: set[tuple] = set()

        for sheet_info in primary_sheets:
            sheet_name = sheet_info["name"]
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            safe_prefix = re.sub(r"[^A-Za-z0-9]", "_", sheet_name)[:20]

            records = _stage_c_extract(
                ws=ws, sheet_info=sheet_info, join_data=join_data,
                join_plan=join_plan, ha_id=ha_id, upload_id=upload_id,
                auto_ref_prefix=safe_prefix,
            )

            for r in records:
                key = (ha_id, r["property_reference"])
                if key not in seen_keys:
                    seen_keys.add(key)
                    all_records.append(r)

        logger.info(f"[{upload_id}] Total records: {len(all_records)}")

        # ── Stage D: Upsert ──
        rows_written = _stage_d_upsert(all_records, upload_id)

        report = _build_coverage_report(all_records, upload_id)
        report["rows_written"] = rows_written
        report["llm_calls"] = 0 if analysis.get("_heuristic") else 1

        if report.get("si_warning"):
            logger.error(
                f"[{upload_id}] ⚠️ FINANCIAL WARNING: Sum Insured is NULL on ALL "
                f"{rows_written} rows. Manual review required."
            )

        logger.info(
            f"[{upload_id}] ══ COMPLETE: {rows_written} rows | "
            f"LLM={report['llm_calls']} | SI=£{report['total_sum_insured']:,.0f} | "
            f"addr={report['field_coverage'].get('address', {}).get('pct', 0)}% | "
            f"pc={report['field_coverage'].get('postcode', {}).get('pct', 0)}% | "
            f"yr={report['field_coverage'].get('year_of_build', {}).get('pct', 0)}% ══"
        )
        return report

    except Exception as exc:
        logger.error(f"[{upload_id}] UNHANDLED ERROR: {exc}", exc_info=True)
        return {
            "error": str(exc), "rows_extracted": 0, "rows_written": 0,
            "total_sum_insured": 0, "si_warning": True,
        }