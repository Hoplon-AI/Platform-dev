# """
# SOV (Schedule of Values) processor: Reads CSV/Excel property schedules from S3 and writes to Silver layer.

# This processor:
# 1. Reads property schedule files (CSV/Excel) from S3 Bronze
# 2. Normalizes column names to standard schema
# 3. Validates data (UPRN format, postcode, required fields)
# 4. Upserts to silver.properties table
# 5. Tracks UPRN lineage for data provenance
# 6. Updates upload_audit with processing status

# Triggered by:
# - Step Functions state machine (for property_schedule dataset type)
# - Or direct invocation for batch processing
# """

# from __future__ import annotations

# import asyncio
# import io
# import json
# import os
# import re
# import uuid
# from datetime import datetime, timezone
# from typing import Any, Dict, List, Optional, Tuple

# import asyncpg
# import boto3
# import pandas as pd

# from infrastructure.storage.s3_config import S3Config
# from infrastructure.storage.upload_service import UploadService
# from backend.core.database.db_pool import DatabasePool


# def _utc_now() -> datetime:
#     return datetime.now(timezone.utc)


# # Column mapping: various input column names -> standard column name
# # Based on schemas/standardized-property-schema.json
# COLUMN_MAPPING: Dict[str, str] = {
#     # UPRN variations
#     "uprn": "uprn",
#     "unique_property_reference": "uprn",
#     "property_reference": "uprn",
#     "unique_property_reference_number": "uprn",

#     # Address variations
#     "address": "address",
#     "full_address": "address",
#     "property_address": "address",
#     "street_address": "address",

#     # Postcode variations
#     "postcode": "postcode",
#     "post_code": "postcode",
#     "zip_code": "postcode",
#     "postal_code": "postcode",

#     # Latitude variations
#     "latitude": "latitude",
#     "lat": "latitude",
#     "y": "latitude",

#     # Longitude variations
#     "longitude": "longitude",
#     "lon": "longitude",
#     "lng": "longitude",
#     "x": "longitude",

#     # Units variations
#     "units": "units",
#     "num_units": "units",
#     "number_of_units": "units",
#     "total_units": "units",

#     # Height variations
#     "height_m": "height_m",
#     "height": "height_m",
#     "building_height": "height_m",
#     "height_metres": "height_m",

#     # Build year variations
#     "build_year": "build_year",
#     "buildyear": "build_year",
#     "year_built": "build_year",
#     "construction_year": "build_year",
#     "built": "build_year",

#     # Construction type variations
#     "construction_type": "construction_type",
#     "construction": "construction_type",
#     "material": "construction_type",
#     "building_type": "construction_type",

#     # Tenure variations
#     "tenure": "tenure",
#     "ownership": "tenure",
#     "tenancy_type": "tenure",

#     # Risk rating variations
#     "risk_rating": "risk_rating",
#     "risk": "risk_rating",
#     "risk_band": "risk_rating",
#     "risk_level": "risk_rating",
#     "riskband": "risk_rating",

#     # Block name variations
#     "block_name": "block_name",
#     "block": "block_name",
#     "block_ref": "block_name",
# }

# # UK postcode regex pattern
# UK_POSTCODE_PATTERN = re.compile(
#     r'^([A-Z]{1,2}[0-9][0-9A-Z]?\s*[0-9][A-Z]{2})$',
#     re.IGNORECASE
# )

# # UPRN pattern (12 digits)
# UPRN_PATTERN = re.compile(r'^\d{12}$')


# class ValidationError:
#     """Container for a validation error."""
#     def __init__(self, row_index: int, field: str, message: str, value: Any = None):
#         self.row_index = row_index
#         self.field = field
#         self.message = message
#         self.value = value

#     def to_dict(self) -> Dict[str, Any]:
#         return {
#             "row_index": self.row_index,
#             "field": self.field,
#             "message": self.message,
#             "value": str(self.value) if self.value is not None else None,
#         }


# async def _get_db_connection(
#     conn: Optional[asyncpg.Connection] = None,
#     pool: Optional[asyncpg.Pool] = None,
# ) -> Tuple[asyncpg.Connection, bool]:
#     """
#     Get a database connection using dependency injection.

#     Args:
#         conn: Optional existing connection (for testing/mocking)
#         pool: Optional connection pool (for dependency injection)

#     Returns:
#         Tuple of (connection, should_release)

#     Raises:
#         RuntimeError: If no connection can be obtained
#     """
#     if conn is not None:
#         return conn, False

#     if pool is not None:
#         return await pool.acquire(), True

#     try:
#         db_pool = DatabasePool.get_pool()
#         return await db_pool.acquire(), True
#     except RuntimeError:
#         host = os.getenv("DB_HOST", "localhost")
#         port = int(os.getenv("DB_PORT", "5432"))
#         database = os.getenv("DB_NAME", "platform_dev")

#         user = os.getenv("DB_USER", "postgres")
#         password = os.getenv("DB_PASSWORD", "postgres")

#         secret_arn = os.getenv("DATABASE_SECRET_ARN")
#         if secret_arn:
#             sm = boto3.client("secretsmanager")
#             resp = sm.get_secret_value(SecretId=secret_arn)
#             secret_str = resp.get("SecretString") or "{}"
#             try:
#                 secret = json.loads(secret_str)
#                 user = secret.get("username", user)
#                 password = secret.get("password", password)
#             except Exception:
#                 pass

#         connection = await asyncpg.connect(
#             host=host,
#             port=port,
#             user=user,
#             password=password,
#             database=database,
#         )
#         return connection, False


# def _parse_s3_key_for_metadata(key: str) -> Dict[str, str]:
#     """
#     Parse S3 key to extract ha_id, submission_id, and file_type.

#     Expected format: ha_id=<ha_id>/bronze/dataset=<file_type>/ingest_date=YYYY-MM-DD/submission_id=<uuid>/...
#     """
#     ha_match = re.search(r"ha_id=([^/]+)/", key)
#     dataset_match = re.search(r"dataset=([^/]+)/", key)
#     submission_match = re.search(r"submission_id=([0-9a-fA-F-]{36})/", key)

#     if not ha_match or not dataset_match or not submission_match:
#         raise ValueError(f"Could not parse S3 key: {key}")

#     return {
#         "ha_id": ha_match.group(1),
#         "file_type": dataset_match.group(1),
#         "submission_id": submission_match.group(1),
#     }


# def _detect_file_type(filename: str) -> str:
#     """Detect file type from filename extension."""
#     lower_name = filename.lower()
#     if lower_name.endswith('.csv'):
#         return 'csv'
#     elif lower_name.endswith(('.xlsx', '.xls')):
#         return 'excel'
#     else:
#         raise ValueError(f"Unsupported file type: {filename}")


# def _parse_sov_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
#     """
#     Parse SOV file (CSV or Excel) into DataFrame.

#     Args:
#         file_bytes: Raw file content
#         filename: Original filename (used for type detection)

#     Returns:
#         pandas DataFrame with the file contents
#     """
#     file_type = _detect_file_type(filename)

#     if file_type == 'csv':
#         # Try different encodings
#         for encoding in ['utf-8', 'latin-1', 'cp1252']:
#             try:
#                 df = pd.read_csv(
#                     io.BytesIO(file_bytes),
#                     encoding=encoding,
#                     dtype=str,  # Read all as strings initially
#                     na_values=['', 'NA', 'N/A', 'NULL', 'null', 'None'],
#                     keep_default_na=True,
#                 )
#                 break
#             except UnicodeDecodeError:
#                 continue
#         else:
#             raise ValueError("Could not decode CSV file with any supported encoding")
#     else:
#         df = pd.read_excel(
#             io.BytesIO(file_bytes),
#             dtype=str,
#             na_values=['', 'NA', 'N/A', 'NULL', 'null', 'None'],
#         )

#     # Strip whitespace from column names
#     df.columns = df.columns.str.strip().str.lower()

#     # Strip whitespace from all string values
#     for col in df.columns:
#         if df[col].dtype == object:
#             df[col] = df[col].str.strip()

#     return df


# def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
#     """
#     Normalize column names to standard schema.

#     Args:
#         df: Input DataFrame

#     Returns:
#         DataFrame with standardized column names
#     """
#     # Create mapping for this DataFrame's columns
#     rename_map = {}
#     for col in df.columns:
#         col_lower = col.lower().strip()
#         if col_lower in COLUMN_MAPPING:
#             standard_name = COLUMN_MAPPING[col_lower]
#             if standard_name not in rename_map.values():  # Avoid duplicates
#                 rename_map[col] = standard_name

#     return df.rename(columns=rename_map)


# def _validate_uprn(uprn: Optional[str]) -> Tuple[bool, Optional[str]]:
#     """
#     Validate UPRN format.

#     Args:
#         uprn: UPRN value to validate

#     Returns:
#         Tuple of (is_valid, cleaned_uprn)
#     """
#     if uprn is None or pd.isna(uprn) or str(uprn).strip() == '':
#         return True, None  # UPRN is optional

#     uprn_str = str(uprn).strip()

#     # UPRN should only contain digits (after stripping whitespace)
#     if not uprn_str.isdigit():
#         return False, uprn_str

#     # UPRN should be up to 12 digits
#     if len(uprn_str) > 12:
#         return False, uprn_str

#     # Pad shorter UPRNs to 12 digits
#     if len(uprn_str) <= 12:
#         return True, uprn_str.zfill(12)

#     return False, uprn_str


# def _validate_postcode(postcode: Optional[str]) -> Tuple[bool, Optional[str]]:
#     """
#     Validate UK postcode format.

#     Args:
#         postcode: Postcode value to validate

#     Returns:
#         Tuple of (is_valid, cleaned_postcode)
#     """
#     if postcode is None or pd.isna(postcode) or str(postcode).strip() == '':
#         return True, None  # Postcode is optional

#     postcode_str = str(postcode).strip().upper()

#     # Remove all whitespace first, then try to format properly
#     postcode_no_space = re.sub(r'\s+', '', postcode_str)

#     # Valid UK postcodes are 5-7 characters without space
#     if len(postcode_no_space) >= 5 and len(postcode_no_space) <= 7:
#         # Inward code is always 3 characters (digit + 2 letters)
#         inward = postcode_no_space[-3:]
#         outward = postcode_no_space[:-3]
#         formatted = f"{outward} {inward}"

#         if UK_POSTCODE_PATTERN.match(formatted):
#             return True, formatted

#     return False, postcode_str


# def _validate_latitude(lat: Optional[str]) -> Tuple[bool, Optional[float]]:
#     """
#     Validate latitude (UK range: approximately 49.0-61.0).

#     Args:
#         lat: Latitude value to validate

#     Returns:
#         Tuple of (is_valid, parsed_latitude)
#     """
#     if lat is None or pd.isna(lat) or str(lat).strip() == '':
#         return True, None

#     try:
#         lat_float = float(lat)
#         # UK approximate range
#         if 49.0 <= lat_float <= 61.0:
#             return True, lat_float
#         return False, lat_float
#     except (ValueError, TypeError):
#         return False, None


# def _validate_longitude(lon: Optional[str]) -> Tuple[bool, Optional[float]]:
#     """
#     Validate longitude (UK range: approximately -8.0 to 2.0).

#     Args:
#         lon: Longitude value to validate

#     Returns:
#         Tuple of (is_valid, parsed_longitude)
#     """
#     if lon is None or pd.isna(lon) or str(lon).strip() == '':
#         return True, None

#     try:
#         lon_float = float(lon)
#         # UK approximate range
#         if -8.0 <= lon_float <= 2.0:
#             return True, lon_float
#         return False, lon_float
#     except (ValueError, TypeError):
#         return False, None


# def _validate_risk_rating(rating: Optional[str]) -> Tuple[bool, Optional[str]]:
#     """
#     Validate risk rating (A-E or null).

#     Args:
#         rating: Risk rating value to validate

#     Returns:
#         Tuple of (is_valid, cleaned_rating)
#     """
#     if rating is None or pd.isna(rating) or str(rating).strip() == '':
#         return True, None

#     rating_str = str(rating).strip().upper()

#     if rating_str in ('A', 'B', 'C', 'D', 'E'):
#         return True, rating_str

#     return False, rating_str


# def _parse_integer(value: Optional[str]) -> Optional[int]:
#     """Parse a value as integer, returning None if not valid."""
#     if value is None or pd.isna(value) or str(value).strip() == '':
#         return None
#     try:
#         return int(float(value))  # Handle "42.0" -> 42
#     except (ValueError, TypeError):
#         return None


# def _parse_float(value: Optional[str]) -> Optional[float]:
#     """Parse a value as float, returning None if not valid."""
#     if value is None or pd.isna(value) or str(value).strip() == '':
#         return None
#     try:
#         return float(value)
#     except (ValueError, TypeError):
#         return None


# def _validate_row(
#     row: pd.Series,
#     row_index: int,
# ) -> Tuple[Dict[str, Any], List[ValidationError]]:
#     """
#     Validate a single row and return cleaned values and errors.

#     Args:
#         row: DataFrame row
#         row_index: Index of the row (for error reporting)

#     Returns:
#         Tuple of (cleaned_values dict, list of validation errors)
#     """
#     errors: List[ValidationError] = []
#     cleaned: Dict[str, Any] = {}

#     # Validate and clean UPRN
#     uprn_valid, uprn_clean = _validate_uprn(row.get('uprn'))
#     if not uprn_valid:
#         errors.append(ValidationError(row_index, 'uprn', 'Invalid UPRN format', row.get('uprn')))
#     cleaned['uprn'] = uprn_clean

#     # Address is required
#     address = row.get('address')
#     if address is None or pd.isna(address) or str(address).strip() == '':
#         errors.append(ValidationError(row_index, 'address', 'Address is required'))
#         cleaned['address'] = None
#     else:
#         cleaned['address'] = str(address).strip()

#     # Validate and clean postcode
#     pc_valid, pc_clean = _validate_postcode(row.get('postcode'))
#     if not pc_valid:
#         errors.append(ValidationError(row_index, 'postcode', 'Invalid UK postcode format', row.get('postcode')))
#     cleaned['postcode'] = pc_clean

#     # Validate and clean latitude
#     lat_valid, lat_clean = _validate_latitude(row.get('latitude'))
#     if not lat_valid:
#         errors.append(ValidationError(row_index, 'latitude', 'Latitude outside UK range (49.0-61.0)', row.get('latitude')))
#         lat_clean = None  # Don't use invalid coordinates
#     cleaned['latitude'] = lat_clean

#     # Validate and clean longitude
#     lon_valid, lon_clean = _validate_longitude(row.get('longitude'))
#     if not lon_valid:
#         errors.append(ValidationError(row_index, 'longitude', 'Longitude outside UK range (-8.0 to 2.0)', row.get('longitude')))
#         lon_clean = None  # Don't use invalid coordinates
#     cleaned['longitude'] = lon_clean

#     # Parse numeric fields
#     cleaned['units'] = _parse_integer(row.get('units'))
#     cleaned['height_m'] = _parse_float(row.get('height_m'))
#     cleaned['build_year'] = _parse_integer(row.get('build_year'))

#     # Validate build year if present
#     if cleaned['build_year'] is not None:
#         current_year = datetime.now().year
#         if cleaned['build_year'] < 1600 or cleaned['build_year'] > current_year + 5:
#             errors.append(ValidationError(
#                 row_index, 'build_year',
#                 f'Build year out of reasonable range (1600-{current_year + 5})',
#                 row.get('build_year')
#             ))

#     # Validate and clean risk rating
#     risk_valid, risk_clean = _validate_risk_rating(row.get('risk_rating'))
#     if not risk_valid:
#         errors.append(ValidationError(row_index, 'risk_rating', 'Invalid risk rating (must be A-E)', row.get('risk_rating')))
#     cleaned['risk_rating'] = risk_clean

#     # String fields (simple cleaning)
#     cleaned['construction_type'] = str(row.get('construction_type')).strip() if row.get('construction_type') and not pd.isna(row.get('construction_type')) else None
#     cleaned['tenure'] = str(row.get('tenure')).strip() if row.get('tenure') and not pd.isna(row.get('tenure')) else None
#     cleaned['block_name'] = str(row.get('block_name')).strip() if row.get('block_name') and not pd.isna(row.get('block_name')) else None

#     # Truncate string fields to fit database constraints
#     if cleaned['construction_type'] and len(cleaned['construction_type']) > 50:
#         cleaned['construction_type'] = cleaned['construction_type'][:50]
#     if cleaned['tenure'] and len(cleaned['tenure']) > 50:
#         cleaned['tenure'] = cleaned['tenure'][:50]
#     if cleaned['risk_rating'] and len(cleaned['risk_rating']) > 2:
#         cleaned['risk_rating'] = cleaned['risk_rating'][:2]

#     return cleaned, errors


# async def _upsert_property(
#     conn: asyncpg.Connection,
#     ha_id: str,
#     upload_id: uuid.UUID,
#     cleaned: Dict[str, Any],
# ) -> uuid.UUID:
#     """
#     Upsert a property record to silver.properties.

#     Uses ON CONFLICT (ha_id, uprn) DO UPDATE for properties with UPRN.
#     Creates new record for properties without UPRN.

#     Args:
#         conn: Database connection
#         ha_id: Housing association ID
#         upload_id: Upload ID for lineage
#         cleaned: Cleaned and validated property data

#     Returns:
#         property_id (UUID)
#     """
#     now = _utc_now().replace(tzinfo=None)

#     if cleaned['uprn']:
#         # Upsert based on ha_id + uprn
#         row = await conn.fetchrow(
#             """
#             INSERT INTO silver.properties (
#                 property_id, ha_id, uprn, address, postcode,
#                 latitude, longitude, units, height_m, build_year,
#                 construction_type, tenure, risk_rating,
#                 created_at, updated_at, metadata
#             )
#             VALUES (
#                 gen_random_uuid(), $1, $2, $3, $4,
#                 $5, $6, $7, $8, $9,
#                 $10, $11, $12,
#                 $13, $14, $15::jsonb
#             )
#             ON CONFLICT (ha_id, uprn) DO UPDATE SET
#                 address = COALESCE(EXCLUDED.address, silver.properties.address),
#                 postcode = COALESCE(EXCLUDED.postcode, silver.properties.postcode),
#                 latitude = COALESCE(EXCLUDED.latitude, silver.properties.latitude),
#                 longitude = COALESCE(EXCLUDED.longitude, silver.properties.longitude),
#                 units = COALESCE(EXCLUDED.units, silver.properties.units),
#                 height_m = COALESCE(EXCLUDED.height_m, silver.properties.height_m),
#                 build_year = COALESCE(EXCLUDED.build_year, silver.properties.build_year),
#                 construction_type = COALESCE(EXCLUDED.construction_type, silver.properties.construction_type),
#                 tenure = COALESCE(EXCLUDED.tenure, silver.properties.tenure),
#                 risk_rating = COALESCE(EXCLUDED.risk_rating, silver.properties.risk_rating),
#                 updated_at = $14,
#                 metadata = COALESCE(silver.properties.metadata, '{}'::jsonb) || $15::jsonb
#             RETURNING property_id
#             """,
#             ha_id,
#             cleaned['uprn'],
#             cleaned['address'],
#             cleaned['postcode'],
#             cleaned['latitude'],
#             cleaned['longitude'],
#             cleaned['units'],
#             cleaned['height_m'],
#             cleaned['build_year'],
#             cleaned['construction_type'],
#             cleaned['tenure'],
#             cleaned['risk_rating'],
#             now,
#             now,
#             json.dumps({"last_upload_id": str(upload_id)}),
#         )
#     else:
#         # No UPRN - create new record (or could match on address, but that's risky)
#         row = await conn.fetchrow(
#             """
#             INSERT INTO silver.properties (
#                 property_id, ha_id, uprn, address, postcode,
#                 latitude, longitude, units, height_m, build_year,
#                 construction_type, tenure, risk_rating,
#                 created_at, updated_at, metadata
#             )
#             VALUES (
#                 gen_random_uuid(), $1, $2, $3, $4,
#                 $5, $6, $7, $8, $9,
#                 $10, $11, $12,
#                 $13, $14, $15::jsonb
#             )
#             RETURNING property_id
#             """,
#             ha_id,
#             cleaned['uprn'],  # None
#             cleaned['address'],
#             cleaned['postcode'],
#             cleaned['latitude'],
#             cleaned['longitude'],
#             cleaned['units'],
#             cleaned['height_m'],
#             cleaned['build_year'],
#             cleaned['construction_type'],
#             cleaned['tenure'],
#             cleaned['risk_rating'],
#             now,
#             now,
#             json.dumps({"last_upload_id": str(upload_id)}),
#         )

#     return row['property_id']


# async def _link_uprn_lineage(
#     conn: asyncpg.Connection,
#     ha_id: str,
#     uprn: str,
#     submission_id: uuid.UUID,
#     property_id: uuid.UUID,
# ) -> None:
#     """
#     Track UPRN lineage - links UPRN to submission and property.

#     Args:
#         conn: Database connection
#         ha_id: Housing association ID
#         uprn: UPRN value
#         submission_id: Upload/submission ID
#         property_id: Property ID in silver.properties
#     """
#     now = _utc_now().replace(tzinfo=None)

#     await conn.execute(
#         """
#         INSERT INTO uprn_lineage_map (
#             uprn, ha_id, submission_id, property_id,
#             first_seen_at, last_updated_at
#         )
#         VALUES ($1, $2, $3, $4, $5, $6)
#         ON CONFLICT (uprn, ha_id, submission_id) DO UPDATE SET
#             property_id = EXCLUDED.property_id,
#             last_updated_at = EXCLUDED.last_updated_at
#         """,
#         uprn,
#         ha_id,
#         submission_id,
#         property_id,
#         now,
#         now,
#     )


# async def _update_upload_audit_sov(
#     conn: asyncpg.Connection,
#     upload_id: uuid.UUID,
#     ha_id: str,
#     status: str,
#     records_processed: int,
#     records_failed: int,
#     validation_errors: List[Dict[str, Any]],
# ) -> None:
#     """
#     Update upload_audit with SOV processing results.

#     Args:
#         conn: Database connection
#         upload_id: Upload ID
#         ha_id: Housing association ID
#         status: Final status ('completed', 'completed_with_warnings', 'failed')
#         records_processed: Number of records successfully processed
#         records_failed: Number of records that failed validation
#         validation_errors: List of validation error details
#     """
#     now = _utc_now().replace(tzinfo=None)

#     metadata = {
#         "sov_processing": {
#             "records_processed": records_processed,
#             "records_failed": records_failed,
#             "validation_errors": validation_errors[:100],  # Limit stored errors
#             "total_validation_errors": len(validation_errors),
#         }
#     }

#     await conn.execute(
#         """
#         UPDATE upload_audit
#         SET
#             status = $3,
#             processing_completed_at = $4,
#             metadata = COALESCE(metadata, '{}'::jsonb) || $5::jsonb
#         WHERE upload_id = $1 AND ha_id = $2
#         """,
#         upload_id,
#         ha_id,
#         status,
#         now,
#         json.dumps(metadata),
#     )


# async def _insert_processing_audit_sov(
#     conn: asyncpg.Connection,
#     *,
#     ha_id: str,
#     upload_id: uuid.UUID,
#     status: str,
#     records_processed: int,
#     records_failed: int,
#     execution_arn: Optional[str] = None,
# ) -> None:
#     """Insert processing audit record for SOV processing."""
#     now = _utc_now().replace(tzinfo=None)

#     await conn.execute(
#         """
#         INSERT INTO processing_audit (
#             processing_id, ha_id, source_type, source_id,
#             target_type, target_id, transformation_type,
#             started_at, completed_at, status, metadata,
#             attempt, max_attempts, last_error, next_attempt_at, retryable, stepfn_execution_arn
#         )
#         VALUES (
#             $1, $2, 'upload', $3,
#             'properties', $4, 'sov_to_silver_v1',
#             $5, $6, $7, $8::jsonb,
#             1, 1, NULL, NULL, false, $9
#         )
#         """,
#         uuid.uuid4(),
#         ha_id,
#         upload_id,
#         uuid.uuid4(),  # Target placeholder
#         now,
#         now,
#         status,
#         json.dumps({
#             "records_processed": records_processed,
#             "records_failed": records_failed,
#         }),
#         execution_arn,
#     )


# async def process_sov_to_silver(
#     event: Dict[str, Any],
#     *,
#     db_conn: Optional[asyncpg.Connection] = None,
#     db_pool: Optional[asyncpg.Pool] = None,
#     upload_service: Optional[UploadService] = None,
# ) -> Dict[str, Any]:
#     """
#     Process SOV (Schedule of Values) CSV/Excel from S3 Bronze to Silver layer.

#     Flow:
#     1. Read file from S3
#     2. Parse CSV/Excel into DataFrame
#     3. Normalize column names (using standardized-property-schema.json)
#     4. Validate data (UPRN format, postcode format, required fields)
#     5. Upsert to silver.properties (ON CONFLICT UPDATE)
#     6. Track UPRN lineage
#     7. Update upload_audit status

#     Event format (from Step Functions):
#     {
#         "bucket": "bucket-name",
#         "key": "ha_id=.../bronze/dataset=property_schedule/.../file=...",
#         "execution_arn": "arn:..."
#     }

#     Returns:
#         Dict with processing status and statistics
#     """
#     # Parse event
#     if "bucket" in event and "key" in event:
#         bucket = event["bucket"]
#         key = event["key"]
#         execution_arn = event.get("execution_arn")
#     elif "Records" in event and event["Records"]:
#         record = event["Records"][0]
#         bucket = record["s3"]["bucket"]["name"]
#         key = record["s3"]["object"]["key"]
#         execution_arn = None
#     else:
#         raise ValueError("Invalid event format: missing bucket/key or Records")

#     # Parse metadata from S3 key
#     metadata = _parse_s3_key_for_metadata(key)
#     ha_id = metadata["ha_id"]
#     upload_id_uuid = uuid.UUID(metadata["submission_id"])
#     file_type = metadata["file_type"]

#     # Extract filename from key
#     filename_match = re.search(r'/file=([^/]+)$', key)
#     filename = filename_match.group(1) if filename_match else key.split('/')[-1]

#     # Setup S3 service
#     if upload_service is None:
#         s3_cfg = S3Config(bucket_name=bucket)
#         upload_service = UploadService(s3_cfg)

#     # Read file from S3
#     try:
#         file_bytes = upload_service.get_file(key)
#     except Exception as e:
#         return {
#             "status": "failed",
#             "reason": "failed_to_read_file",
#             "error": str(e),
#             "key": key,
#         }

#     # Parse file
#     try:
#         df = _parse_sov_file(file_bytes, filename)
#     except Exception as e:
#         return {
#             "status": "failed",
#             "reason": "failed_to_parse_file",
#             "error": str(e),
#             "key": key,
#         }

#     if df.empty:
#         return {
#             "status": "failed",
#             "reason": "empty_file",
#             "key": key,
#         }

#     # Normalize columns
#     df = _normalize_columns(df)

#     # Check if we have address column (required)
#     if 'address' not in df.columns:
#         return {
#             "status": "failed",
#             "reason": "missing_required_column",
#             "error": "Address column not found. Expected one of: address, full_address, property_address, street_address",
#             "key": key,
#             "columns_found": list(df.columns),
#         }

#     # Database connection
#     conn, should_release = await _get_db_connection(conn=db_conn, pool=db_pool)
#     should_close_conn = not should_release and db_conn is None

#     try:
#         # Set tenant context
#         await conn.execute("SELECT set_config('app.current_ha_id', $1, true)", ha_id)

#         # Process each row
#         records_processed = 0
#         records_failed = 0
#         all_errors: List[ValidationError] = []
#         property_ids: List[uuid.UUID] = []

#         for idx, row in df.iterrows():
#             cleaned, errors = _validate_row(row, int(idx))

#             # Skip rows without valid address
#             if cleaned['address'] is None:
#                 records_failed += 1
#                 all_errors.extend(errors)
#                 continue

#             # Record non-critical errors but continue processing
#             all_errors.extend(errors)

#             try:
#                 property_id = await _upsert_property(
#                     conn, ha_id, upload_id_uuid, cleaned
#                 )
#                 property_ids.append(property_id)

#                 # Track UPRN lineage if UPRN is present
#                 if cleaned['uprn']:
#                     await _link_uprn_lineage(
#                         conn, ha_id, cleaned['uprn'], upload_id_uuid, property_id
#                     )

#                 records_processed += 1

#             except Exception as e:
#                 records_failed += 1
#                 all_errors.append(ValidationError(
#                     int(idx), 'database', f'Database error: {str(e)}'
#                 ))

#         # Determine final status
#         if records_processed == 0:
#             status = 'failed'
#         elif records_failed > 0 or len(all_errors) > 0:
#             status = 'completed_with_warnings'
#         else:
#             status = 'completed'

#         # Update upload_audit
#         await _update_upload_audit_sov(
#             conn,
#             upload_id_uuid,
#             ha_id,
#             status,
#             records_processed,
#             records_failed,
#             [e.to_dict() for e in all_errors],
#         )

#         # Insert processing audit
#         await _insert_processing_audit_sov(
#             conn,
#             ha_id=ha_id,
#             upload_id=upload_id_uuid,
#             status=status,
#             records_processed=records_processed,
#             records_failed=records_failed,
#             execution_arn=execution_arn,
#         )

#         return {
#             "status": status,
#             "ha_id": ha_id,
#             "upload_id": str(upload_id_uuid),
#             "records_total": len(df),
#             "records_processed": records_processed,
#             "records_failed": records_failed,
#             "validation_errors_count": len(all_errors),
#             "property_ids": [str(pid) for pid in property_ids[:10]],  # First 10
#             # Metadata structure for Step Functions compatibility
#             # The Silver processor will ignore this key (not features.json)
#             "metadata": {
#                 "features_s3_key": key,  # Use original key - silver processor will ignore non-features.json
#                 "processing_type": "sov_to_silver",
#                 "silver_processing_complete": True,  # Marker that silver processing is already done
#             },
#         }

#     except Exception as e:
#         return {
#             "status": "failed",
#             "error": str(e),
#             "ha_id": ha_id,
#             "upload_id": str(upload_id_uuid),
#         }

#     finally:
#         if should_release:
#             if db_pool is not None:
#                 await db_pool.release(conn)
#             else:
#                 try:
#                     pool = DatabasePool.get_pool()
#                     await pool.release(conn)
#                 except RuntimeError:
#                     await conn.close()
#         elif should_close_conn:
#             await conn.close()


# def handler(event: Dict[str, Any], context: Any = None) -> Dict[str, Any]:
#     """Lambda-compatible synchronous handler."""
#     return asyncio.run(process_sov_to_silver(event))


# def is_sov_type(file_type: str, filename: str) -> bool:
#     """
#     Check if a file is a property schedule (SOV) file.

#     Args:
#         file_type: Dataset type from S3 key (e.g., 'property_schedule')
#         filename: Original filename

#     Returns:
#         True if this is a property schedule file
#     """
#     # Check dataset type
#     if file_type in ('property_schedule', 'sov', 'schedule_of_values'):
#         return True

#     # Check file extension
#     lower_name = filename.lower()
#     if lower_name.endswith(('.csv', '.xlsx', '.xls')):
#         # Could be a property schedule based on extension alone
#         # but we require explicit dataset type for safety
#         return file_type in ('property_schedule', 'sov', 'schedule_of_values')

#     return False


"""
sov_processor.py
----------------
Processes SOV (Schedule of Values) Excel files in Doc A / Example 11 format.

Expected input format (Stock Listing sheet):
  Col 1:  Client Name
  Col 2:  Start Date
  Col 3:  End Date
  Col 4:  Policy Reference
  Col 5:  Product Type
  Col 6:  Property Reference       → property_reference
  Col 7:  Block Reference          → block_reference
  Col 8:  Occupancy Type           → occupancy_type
  Col 9:  Deductible               (skipped - insurer field)
  Col 10: Flood Deductible         (skipped)
  Col 11: Storm Deductible         (skipped)
  Col 12: Basis of Deductible      (skipped)
  Col 13: Address 1                → address  (may be full address)
  Col 14: Address 2                → address_2
  Col 15: Address 3                → address_3
  Col 16: Postcode                 → postcode
  Col 17: Number of Units          → units
  Col 18: Sum Insured              → sum_insured
         (or: 2026 Sum Insured for year-versioned files)
  Col 19: Sum Insured Type         → sum_insured_type
  Col 20: Property Type            → property_type
  Col 21: Avid Property Type       → avid_property_type
  Col 22: Wall Construction        → wall_construction
  Col 23: Roof Construction        → roof_construction
  Col 24: Floor Construction       → floor_construction
  Col 25: Year of Build            → build_year
  Col 26: Age Banding              → age_banding (also derived from build_year)
  Col 27: Number of Bedrooms       → num_bedrooms
  Col 28: Number of Storeys        → storeys
  Col 29: Basement location        → basement
  Col 30: Listed building          → is_listed
  Col 31: Security Features        → security_features
  Col 32: Fire Protection          → fire_protection
  Col 33: Alarms                   → alarms
  Col 34: Flood insured            → flood_insured
  Col 35: Storm insured            → storm_insured

After DB upsert, calls UPRN matching for each property postcode.
"""

import io
import re
import logging
from datetime import datetime
from typing import Optional
import openpyxl
import asyncpg

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column name normaliser
# Maps the many variations we see in real SOV files → our internal key
# ---------------------------------------------------------------------------
COLUMN_MAP = {
    # Property identity
    "property reference":   "property_reference",
    "property ref":         "property_reference",
    "prop ref":             "property_reference",
    "reference":            "property_reference",
    "ref":                  "property_reference",

    "block reference":      "block_reference",
    "block ref":            "block_reference",
    "block":                "block_reference",

    "occupancy type":       "occupancy_type",
    "occupancy":            "occupancy_type",
    "tenure":               "occupancy_type",

    # Address
    "address 1":            "address",
    "address1":             "address",
    "address line 1":       "address",
    "full address":         "address",
    "address":              "address",

    "address 2":            "address_2",
    "address2":             "address_2",
    "address line 2":       "address_2",

    "address 3":            "address_3",
    "address3":             "address_3",
    "address line 3":       "address_3",

    "postcode":             "postcode",
    "post code":            "postcode",
    "post_code":            "postcode",
    "pc":                   "postcode",

    # Financial
    "sum insured":          "sum_insured",
    "2026 sum insured":     "sum_insured",
    "2025 sum insured":     "sum_insured_prev",   # keep but prefer 2026
    "sum insured type":     "sum_insured_type",
    "2026 sum insured type":"sum_insured_type",
    "sum insured type ":    "sum_insured_type",

    # Units / bedrooms / storeys
    "number of units":      "units",
    "number of units ":     "units",
    "no of units":          "units",
    "units":                "units",

    "number of bedrooms":   "num_bedrooms",
    "number of bedrooms ":  "num_bedrooms",
    "no of bedrooms":       "num_bedrooms",
    "bedrooms":             "num_bedrooms",
    "beds":                 "num_bedrooms",

    "number of storeys":    "storeys",
    "no of storeys":        "storeys",
    "storeys":              "storeys",
    "number of stories":    "storeys",
    "stories":              "storeys",
    "floors":               "storeys",

    # Property classification
    "property type":        "property_type",
    "prop type":            "property_type",

    "avid property type":   "avid_property_type",
    "avid property type ":  "avid_property_type",

    # Construction
    "wall construction":    "wall_construction",
    "wall construction ":   "wall_construction",
    "wall":                 "wall_construction",

    "roof construction":    "roof_construction",
    "roof":                 "roof_construction",

    "floor construction":   "floor_construction",
    "floor construction ":  "floor_construction",
    "floor":                "floor_construction",

    # Build age
    "year of build":        "build_year",
    "year built":           "build_year",
    "year of construction": "build_year",
    "construction date":    "build_year",
    "date of build":        "build_year",

    "age banding":          "age_banding",
    "age banding ":         "age_banding",
    "age band":             "age_banding",

    # Flags
    "basement location":    "basement",
    "basement":             "basement",
    "basement flats":       "basement",

    "listed building (if blank = not listed) ": "is_listed",
    "listed building":      "is_listed",
    "listed":               "is_listed",

    "security features":    "security_features",
    "security features ":   "security_features",
    "security":             "security_features",

    "fire protection":      "fire_protection",
    "fire protection ":     "fire_protection",

    "alarms":               "alarms",
    "alarms ":              "alarms",

    "flood insured":        "flood_insured",
    "flood insured ":       "flood_insured",

    "storm insured":        "storm_insured",
    "storm insured ":       "storm_insured",
}

# Fields we actively skip (insurer-filled, not from HA)
SKIP_COLUMNS = {
    "client name", "start date", "end date", "policy reference",
    "product type", "deductible", "flood deductible", "storm deductible",
    "basis of deductible (eel/sec)", "basis of deductible",
    "jba if over 10 and gf exposure",
    "2025 sum insured type",
}

# Age banding derivation
AGE_BANDS = [
    (0,    1919, "Pre 1919"),
    (1919, 1944, "1919-1944"),
    (1945, 1964, "1945-1964"),
    (1965, 1980, "1965-1980"),
    (1981, 2000, "1981-2000"),
    (2001, 9999, "Post 2000"),
]

POSTCODE_RE = re.compile(
    r'^[A-Z]{1,2}[0-9][0-9A-Z]?\s?[0-9][A-Z]{2}$',
    re.IGNORECASE
)

# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

async def process_sov_to_silver(
    file_bytes: bytes,
    ha_id: str,
    submission_id: str,
    upload_id: str,
    db_pool,
) -> dict:
    """
    Main entry point. Called by the upload router after S3 upload.

    Returns:
        {
            "status": "completed" | "completed_with_warnings" | "failed",
            "rows_processed": int,
            "rows_inserted": int,
            "rows_updated": int,
            "rows_failed": int,
            "warnings": [str],
            "errors": [str],
        }
    """
    result = {
        "status": "failed",
        "rows_processed": 0,
        "rows_inserted": 0,
        "rows_updated": 0,
        "rows_failed": 0,
        "warnings": [],
        "errors": [],
    }

    try:
        # 1. Open workbook
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,   # returns cached formula results, not formula strings
        )

        # 2. Find the right sheet
        sheet = _select_sheet(wb)
        if sheet is None:
            result["errors"].append("No property data sheet found in workbook.")
            return result
        logger.info(f"[{submission_id}] Using sheet: '{sheet.title}'")

        # 3. Find header row and build column index map
        header_row_idx, col_map = _find_headers(sheet)
        if not col_map:
            result["errors"].append("Could not identify column headers in sheet.")
            return result
        logger.info(f"[{submission_id}] Headers found on row {header_row_idx + 1}. "
                    f"Mapped columns: {list(col_map.keys())}")

        # 4. Validate we have minimum required columns
        required = {"address", "postcode"}
        missing = required - set(col_map.keys())
        if missing:
            result["errors"].append(f"Missing required columns: {missing}")
            return result

        # 5. Process rows
        async with db_pool.acquire() as conn:
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_idx <= header_row_idx:
                    continue

                # Skip empty rows
                if all(c is None or str(c).strip() == "" for c in row):
                    continue

                result["rows_processed"] += 1

                # Extract values using column map
                raw = _extract_row(row, col_map)

                # Clean and validate
                cleaned, row_warnings, row_errors = _clean_row(raw)

                if row_errors:
                    result["rows_failed"] += 1
                    result["errors"].extend(
                        [f"Row {row_idx + 1}: {e}" for e in row_errors]
                    )
                    continue

                result["warnings"].extend(
                    [f"Row {row_idx + 1}: {w}" for w in row_warnings]
                )

                # Upsert to silver.properties
                action = await _upsert_property(conn, ha_id, submission_id, cleaned)
                if action == "insert":
                    result["rows_inserted"] += 1
                else:
                    result["rows_updated"] += 1

                # UPRN matching (best-effort, non-blocking)
                if cleaned.get("postcode") and not cleaned.get("uprn"):
                    await _match_uprn(conn, ha_id, cleaned["postcode"],
                                      cleaned.get("_property_id"))

        result["status"] = (
            "completed_with_warnings"
            if result["warnings"] or result["rows_failed"] > 0
            else "completed"
        )

    except Exception as e:
        logger.exception(f"[{submission_id}] SOV processing failed: {e}")
        result["errors"].append(str(e))

    finally:
        await _update_audit(db_pool, upload_id, result)

    return result


# ---------------------------------------------------------------------------
# Sheet selection
# ---------------------------------------------------------------------------

GOOD_SHEET_SIGNALS = [
    "stock", "listing", "properties", "assets", "dwellings",
    "rented", "schedule", "units", "ews", "sov", "property",
]
BAD_SHEET_SIGNALS = [
    "summary", "broker", "fee", "indexation", "wages", "pivot",
    "notes", "validation", "workings", "voids", "lookup",
    "sheet1", "sheet2", "sheet3",
]


def _select_sheet(wb):
    """Score all sheets and return the one most likely to contain property data."""
    best_sheet = None
    best_score = -999

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        score = _score_sheet(ws, sheet_name)
        logger.debug(f"  Sheet '{sheet_name}': score={score}")
        if score > best_score:
            best_score = score
            best_sheet = ws

    return best_sheet if best_score > 0 else None


def _score_sheet(ws, sheet_name: str) -> int:
    score = 0
    name_lower = sheet_name.lower().strip()

    for signal in GOOD_SHEET_SIGNALS:
        if signal in name_lower:
            score += 3
    for signal in BAD_SHEET_SIGNALS:
        if signal in name_lower:
            score -= 2

    # Read first 10 rows for scoring
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 10:
            break
        rows.append(row)

    if not rows:
        return -10

    # Postcode column name signal
    for cell in rows[0]:
        if cell and "post" in str(cell).lower():
            score += 5

    # Postcode value signal in data rows
    postcode_hits = 0
    for row in rows[1:]:
        for cell in row:
            val = str(cell).strip().upper() if cell else ""
            if POSTCODE_RE.match(val):
                postcode_hits += 1
                break
    score += min(postcode_hits * 2, 8)

    # Column count sweet spot
    col_count = sum(1 for c in rows[0] if c is not None)
    if 8 <= col_count <= 45:
        score += 2

    # Row count
    row_count = sum(1 for r in rows if any(c is not None for c in r))
    if row_count >= 5:
        score += 2

    return score


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _find_headers(ws):
    """
    Scans the first 10 rows to find the header row.
    Returns (header_row_index, {internal_key: col_index}).
    """
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx >= 10:
            break

        # A header row has mostly non-None string values
        text_cells = [
            c for c in row
            if c is not None and isinstance(c, str) and c.strip()
        ]
        total_cells = sum(1 for c in row if c is not None)

        if total_cells == 0:
            continue

        # If >60% of non-null cells are strings → candidate header row
        if len(text_cells) / total_cells >= 0.6 and len(text_cells) >= 3:
            col_map = {}
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                key = COLUMN_MAP.get(str(cell).strip().lower())
                if key and key not in ("sum_insured_prev",):
                    # If we already have sum_insured (from a prior col),
                    # don't overwrite with the older year version
                    if key == "sum_insured" and "sum_insured" in col_map:
                        continue
                    col_map[key] = col_idx

            # Must have at minimum address + postcode to be valid
            if "address" in col_map and "postcode" in col_map:
                return row_idx, col_map

    return 0, {}


# ---------------------------------------------------------------------------
# Row extraction
# ---------------------------------------------------------------------------

def _extract_row(row: tuple, col_map: dict) -> dict:
    """Pull values from the row using the column index map."""
    raw = {}
    for key, col_idx in col_map.items():
        if col_idx < len(row):
            val = row[col_idx]
            raw[key] = val
    return raw


# ---------------------------------------------------------------------------
# Row cleaning and validation
# ---------------------------------------------------------------------------

def _clean_row(raw: dict) -> tuple[dict, list, list]:
    """
    Clean, type-cast and validate a raw row dict.
    Returns (cleaned_dict, warnings, errors).
    errors → row is skipped
    warnings → row is saved with a note
    """
    cleaned = {}
    warnings = []
    errors = []

    # --- Address ---
    addr_parts = [
        _str(raw.get("address")),
        _str(raw.get("address_2")),
        _str(raw.get("address_3")),
    ]
    addr_parts = [p for p in addr_parts if p]
    if not addr_parts:
        errors.append("No address found")
        return cleaned, warnings, errors

    cleaned["address"] = ", ".join(addr_parts)
    cleaned["address_2"] = _str(raw.get("address_2"))
    cleaned["address_3"] = _str(raw.get("address_3"))

    # --- Postcode ---
    pc = _str(raw.get("postcode"))
    if not pc:
        # Try to extract from address
        pc = _extract_postcode_from_address(cleaned["address"])
    if not pc:
        errors.append("No postcode found")
        return cleaned, warnings, errors

    pc = re.sub(r'\s+', ' ', pc.strip().upper())
    if not POSTCODE_RE.match(pc):
        warnings.append(f"Postcode '{pc}' may be invalid")
    cleaned["postcode"] = pc

    # --- String fields ---
    for field in [
        "property_reference", "block_reference", "occupancy_type",
        "sum_insured_type", "avid_property_type",
        "wall_construction", "roof_construction", "floor_construction",
        "security_features", "fire_protection", "alarms",
        "property_type",
    ]:
        cleaned[field] = _str(raw.get(field))

    # --- Numeric fields ---
    cleaned["sum_insured"] = _decimal(raw.get("sum_insured"))
    if cleaned["sum_insured"] is None:
        warnings.append("sum_insured is missing or non-numeric")

    cleaned["units"] = _int(raw.get("units"))
    cleaned["num_bedrooms"] = _int(raw.get("num_bedrooms"))
    cleaned["storeys"] = _int(raw.get("storeys"))

    # --- Build year ---
    build_year = _parse_build_year(raw.get("build_year"))
    cleaned["build_year"] = build_year

    # --- Age banding ---
    age_band = _str(raw.get("age_banding"))
    if not age_band and build_year:
        age_band = _derive_age_band(build_year)
    cleaned["age_banding"] = age_band

    # --- Boolean fields ---
    cleaned["basement"] = _bool(raw.get("basement"))
    cleaned["is_listed"] = _bool(raw.get("is_listed"))
    cleaned["flood_insured"] = _bool(raw.get("flood_insured"))
    cleaned["storm_insured"] = _bool(raw.get("storm_insured"))

    return cleaned, warnings, errors


def _str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("none", "nan", "n/a", "-") else None


def _decimal(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = re.sub(r'[£,\s]', '', str(val))
    try:
        return float(s)
    except ValueError:
        return None


def _int(val) -> Optional[int]:
    v = _decimal(val)
    return int(v) if v is not None else None


def _bool(val) -> Optional[bool]:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("yes", "true", "1", "y"):
        return True
    if s in ("no", "false", "0", "n", ""):
        return False
    return None


def _parse_build_year(val) -> Optional[int]:
    if val is None:
        return None
    # Handle datetime objects (openpyxl returns dates as datetime)
    if isinstance(val, datetime):
        return val.year
    s = str(val).strip()
    # Extract first 4-digit year from strings like "1965 TO 1982" or "1900's"
    match = re.search(r'\b(1[89]\d{2}|20[012]\d)\b', s)
    if match:
        return int(match.group(1))
    return None


def _derive_age_band(year: int) -> str:
    for low, high, label in AGE_BANDS:
        if low <= year <= high:
            return label
    return None


def _extract_postcode_from_address(address: str) -> Optional[str]:
    """Last-resort: try to find a postcode embedded in the address string."""
    match = re.search(
        r'\b([A-Z]{1,2}[0-9][0-9A-Z]?\s?[0-9][A-Z]{2})\b',
        address.upper()
    )
    return match.group(1) if match else None


# ---------------------------------------------------------------------------
# DB upsert
# ---------------------------------------------------------------------------

async def _upsert_property(
    conn: asyncpg.Connection,
    ha_id: str,
    submission_id: str,
    cleaned: dict,
) -> str:
    """
    Upserts a property row into silver.properties.
    Returns "insert" or "update".
    """
    import uuid
    property_id = str(uuid.uuid4())

    sql = """
    INSERT INTO silver.properties (
        property_id,
        ha_id,
        submission_id,
        property_reference,
        block_reference,
        occupancy_type,
        address,
        address_2,
        address_3,
        postcode,
        units,
        sum_insured,
        sum_insured_type,
        property_type,
        avid_property_type,
        wall_construction,
        roof_construction,
        floor_construction,
        build_year,
        age_banding,
        num_bedrooms,
        storeys,
        basement,
        is_listed,
        security_features,
        fire_protection,
        alarms,
        flood_insured,
        storm_insured,
        created_at,
        updated_at
    ) VALUES (
        $1,  $2,  $3,  $4,  $5,  $6,  $7,  $8,  $9,  $10,
        $11, $12, $13, $14, $15, $16, $17, $18, $19, $20,
        $21, $22, $23, $24, $25, $26, $27, $28, $29,
        NOW(), NOW()
    )
    ON CONFLICT (ha_id, property_reference)
    DO UPDATE SET
        submission_id     = EXCLUDED.submission_id,
        property_reference= EXCLUDED.property_reference,
        block_reference   = EXCLUDED.block_reference,
        occupancy_type    = EXCLUDED.occupancy_type,
        address_2         = EXCLUDED.address_2,
        address_3         = EXCLUDED.address_3,
        units             = EXCLUDED.units,
        sum_insured       = EXCLUDED.sum_insured,
        sum_insured_type  = EXCLUDED.sum_insured_type,
        property_type     = EXCLUDED.property_type,
        avid_property_type= EXCLUDED.avid_property_type,
        wall_construction = EXCLUDED.wall_construction,
        roof_construction = EXCLUDED.roof_construction,
        floor_construction= EXCLUDED.floor_construction,
        build_year        = EXCLUDED.build_year,
        age_banding       = EXCLUDED.age_banding,
        num_bedrooms      = EXCLUDED.num_bedrooms,
        storeys           = EXCLUDED.storeys,
        basement          = EXCLUDED.basement,
        is_listed         = EXCLUDED.is_listed,
        security_features = EXCLUDED.security_features,
        fire_protection   = EXCLUDED.fire_protection,
        alarms            = EXCLUDED.alarms,
        flood_insured     = EXCLUDED.flood_insured,
        storm_insured     = EXCLUDED.storm_insured,
        updated_at        = NOW()
    RETURNING (xmax = 0) AS is_insert
    """

    row = await conn.fetchrow(
        sql,
        property_id,
        ha_id,
        submission_id,
        cleaned.get("property_reference"),
        cleaned.get("block_reference"),
        cleaned.get("occupancy_type"),
        cleaned["address"],
        cleaned.get("address_2"),
        cleaned.get("address_3"),
        cleaned["postcode"],
        cleaned.get("units"),
        cleaned.get("sum_insured"),
        cleaned.get("sum_insured_type"),
        cleaned.get("property_type"),
        cleaned.get("avid_property_type"),
        cleaned.get("wall_construction"),
        cleaned.get("roof_construction"),
        cleaned.get("floor_construction"),
        cleaned.get("build_year"),
        cleaned.get("age_banding"),
        cleaned.get("num_bedrooms"),
        cleaned.get("storeys"),
        cleaned.get("basement"),
        cleaned.get("is_listed"),
        cleaned.get("security_features"),
        cleaned.get("fire_protection"),
        cleaned.get("alarms"),
        cleaned.get("flood_insured"),
        cleaned.get("storm_insured"),
    )

    # Store property_id for UPRN matching
    if row:
        cleaned["_property_id"] = property_id
        return "insert" if row["is_insert"] else "update"

    return "update"


# ---------------------------------------------------------------------------
# UPRN matching
# ---------------------------------------------------------------------------

async def _match_uprn(
    conn: asyncpg.Connection,
    ha_id: str,
    postcode: str,
    property_id: Optional[str],
):
    """
    Calls the local PostGIS UPRN lookup for the postcode.
    Updates silver.properties with lat/lng if found.
    Best-effort: errors are logged but do not fail the row.
    """
    try:
        # Query the local uprn_points / postcode_centroids table
        # Same query used by geo/repository.py UPRNRepository
        sql = """
        SELECT
            uprn,
            ST_Y(geom::geometry) AS latitude,
            ST_X(geom::geometry) AS longitude
        FROM uprn_points
        WHERE postcode = $1
        LIMIT 10
        """
        rows = await conn.fetch(sql, postcode.replace(" ", "").upper())

        if not rows:
            # Fall back to postcode centroid
            sql_centroid = """
            SELECT
                NULL AS uprn,
                ST_Y(geom::geometry) AS latitude,
                ST_X(geom::geometry) AS longitude
            FROM postcode_centroids
            WHERE postcode = $1
            LIMIT 1
            """
            rows = await conn.fetch(sql_centroid, postcode.replace(" ", "").upper())

        if rows and property_id:
            best = rows[0]
            uprn_val = str(best["uprn"]) if best["uprn"] else None
            confidence = 90 if len(rows) == 1 else max(10, 90 - len(rows) * 10)

            await conn.execute(
                """
                UPDATE silver.properties
                SET
                    uprn       = $1,
                    latitude   = $2,
                    longitude  = $3,
                    updated_at = NOW()
                WHERE property_id = $4
                  AND ha_id = $5
                  AND (uprn IS NULL OR uprn = '')
                """,
                uprn_val,
                float(best["latitude"]),
                float(best["longitude"]),
                property_id,
                ha_id,
            )
            logger.debug(
                f"UPRN match: {postcode} → {uprn_val} "
                f"({confidence}% confidence, {len(rows)} candidates)"
            )

    except Exception as e:
        logger.warning(f"UPRN match failed for {postcode}: {e}")


# ---------------------------------------------------------------------------
# Audit update
# ---------------------------------------------------------------------------

async def _update_audit(db_pool, upload_id: str, result: dict):
    try:
        async with db_pool.acquire() as conn:
            await conn.execute(
                """
                UPDATE upload_audit
                SET
                    status         = $1,
                    rows_processed = $2,
                    rows_inserted  = $3,
                    rows_failed    = $4,
                    error_details  = $5,
                    completed_at   = NOW()
                WHERE upload_id = $6
                """,
                result["status"],
                result["rows_processed"],
                result["rows_inserted"],
                result["rows_failed"],
                "\n".join(result["errors"] + result["warnings"]) or None,
                upload_id,
            )
    except Exception as e:
        logger.error(f"Failed to update audit for {upload_id}: {e}")