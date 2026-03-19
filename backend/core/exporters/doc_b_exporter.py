"""
doc_b_exporter.py
-----------------
Generates Doc B (High Value Building) Excel file from silver.properties.

Doc B = one row per BLOCK (not per unit), 64 columns.
This exporter fills:
  - Section 2: General Property Details  (Q9–Q27)  ← from silver.properties
  - Section 3: Construction              (Q28–Q31) ← from silver.properties
  - Section 4: Fire Risk Management      (Q32–Q40) ← from silver.fra_features
  - Section 5: EWS / Cladding            (Q41–Q58) ← from silver.fraew_features

Insurer-filled sections (Q1–Q8, Q59–Q64) are left blank.

Usage:
    from doc_b_exporter import generate_doc_b
    excel_bytes = await generate_doc_b(db_pool, ha_id, ha_name)
"""

import io
import logging
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Doc B column definitions — all 64 questions
# source: "sov" = filled from SOV / silver.properties
#         "insurer" = filled by insurer
#         "fra" = filled from FRA document
#         "fraew" = filled from FRAEW document
# ---------------------------------------------------------------------------
DOC_B_COLUMNS = [
    # Section 1 — Policy Information (Q1–Q8) — insurer
    ("Q1",  "Policy number",                                "insurer",  None),
    ("Q2",  "Cover Commencement Date",                      "insurer",  None),
    ("Q3",  "Single location or portfolio?",                "insurer",  None),
    ("Q4",  "Insured Type",                                 "insurer",  None),
    ("Q5",  "Insured Name",                                 "insurer",  None),
    ("Q6",  "Managing Agent?",                              "insurer",  None),
    ("Q7",  "GWP for building",                             "insurer",  None),
    ("Q8",  "Fire Deductible",                              "insurer",  None),

    # Section 2 — General Property Details (Q9–Q27) — from SOV
    ("Q9",  "Block Name/Reference",                         "sov",      "block_reference"),
    ("Q10", "Address Line 1",                               "sov",      "address"),
    ("Q11", "Address Line 2",                               "sov",      "address_2"),
    ("Q12", "Postcode",                                     "sov",      "postcode"),
    ("Q13", "Sum Insured",                                  "sov",      "sum_insured"),
    ("Q14", "LOR/AA/ICOW",                                  "sov",      "lor_value"),
    ("Q15", "Total Insured Value",                          "sov",      "total_insured_value"),
    ("Q16", "Storeys above ground",                         "sov",      "storeys"),
    ("Q17", "Height (metres)",                              "sov",      "height_max_m"),
    ("Q18", "Basement flats present",                       "sov",      "basement"),
    ("Q19", "Property Type",                                "sov",      "property_type"),
    ("Q20", "Occupancy Type",                               "sov",      "occupancy_type"),
    ("Q21", "Is block unoccupied?",                         "sov",      None),
    ("Q22", "Number of units",                              "sov",      "units"),
    ("Q23", "Date of build",                                "sov",      "build_year"),
    ("Q24", "Date of refurb",                               "sov",      None),
    ("Q25", "Refurbishment ongoing?",                       "sov",      None),
    ("Q26", "Listed building?",                             "sov",      "is_listed"),
    ("Q27", "Asylum seekers present?",                      "sov",      None),

    # Section 3 — Construction (Q28–Q31) — from SOV
    ("Q28", "Wall Construction",                            "sov",      "wall_construction"),
    ("Q29", "Floor construction",                           "sov",      "floor_construction"),
    ("Q30", "Roof construction",                            "sov",      "roof_construction"),
    ("Q31", "Timber Framed",                                "sov",      None),

    # Section 4 — Fire Risk Management (Q32–Q40) — from FRA
    ("Q32", "Date of last FRA",                             "fra",      "fra_date"),
    ("Q33", "FRA Report Shared with Insurers?",             "fra",      None),
    ("Q34", "Sprinklers in common parts?",                  "fra",      "sprinklers_common"),
    ("Q35", "Sprinklers in Flats?",                         "fra",      None),
    ("Q36", "Fire Alarms in common parts?",                 "fra",      "fire_alarms_common"),
    ("Q37", "Fire Alarms in Flats?",                        "fra",      None),
    ("Q38", "Waking watch in place?",                       "fra",      None),
    ("Q39", "Other fire protection precautions?",           "fra",      "other_fire_protection"),
    ("Q40", "Security measures in place?",                  "fra",      None),

    # Section 5 — EWS / Cladding (Q41–Q58) — from FRAEW
    ("Q41", "Combustible materials present?",               "fraew",    "combustible_materials"),
    ("Q42", "Which part contains combustibles?",            "fraew",    None),
    ("Q43", "Details of combustibles",                      "fraew",    None),
    ("Q44", "Is remediation work required?",                "fraew",    "remediation_required"),
    ("Q45", "Reasons if no remediation",                    "fraew",    None),
    ("Q46", "Remediation status",                           "fraew",    None),
    ("Q47", "Remediation details",                          "fraew",    "interim_measures_detail"),
    ("Q48", "Remediation completion date",                  "fraew",    None),
    ("Q49", "Cladding present?",                            "fraew",    "cladding_present"),
    ("Q50", "% external walls covered",                     "fraew",    None),
    ("Q51", "Type of cladding",                             "fraew",    None),
    ("Q52", "Cladding combustibility rating (A-E)",         "fraew",    None),
    ("Q53", "Insulation Type",                              "fraew",    "insulation_type"),
    ("Q54", "Insulation Combustibility Rating (A-E)",       "fraew",    None),
    ("Q55", "Date EWS fitted",                              "fraew",    None),
    ("Q56", "Does EWS require remediation?",                "fraew",    "ews_remediation_required"),
    ("Q57", "EWS remediation plan",                         "fraew",    None),
    ("Q58", "EWS remediation date",                         "fraew",    None),

    # Section 6 — Claims (Q59–Q64) — insurer
    ("Q59", "Number of fire claims",                        "insurer",  None),
    ("Q60", "Incurred fire claims amount",                  "insurer",  None),
    ("Q61", "Number of EOW claims",                         "insurer",  None),
    ("Q62", "Incurred EOW claims amount",                   "insurer",  None),
    ("Q63", "Incurred all claims amount",                   "insurer",  None),
    ("Q64", "Additional Comments",                          "insurer",  None),
]

# Colour coding by source
FILLS = {
    "insurer": PatternFill("solid", fgColor="D6E4F0"),  # light blue
    "sov":     PatternFill("solid", fgColor="FFFFFF"),  # white
    "fra":     PatternFill("solid", fgColor="FFF2CC"),  # light yellow
    "fraew":   PatternFill("solid", fgColor="FCE4D6"),  # light orange
}
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
NORMAL_FONT  = Font(name="Calibri", size=9)
THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
ALT_ROW_FILL = PatternFill("solid", fgColor="F5F5F5")


# ---------------------------------------------------------------------------
# Public function
# ---------------------------------------------------------------------------

async def generate_doc_b(
    db_pool,
    ha_id: str,
    ha_name: str,
    portfolio_id: Optional[str] = None,
) -> bytes:
    """
    Query silver.properties grouped by block_reference and generate Doc B.
    Rows with no block_reference are treated as standalone blocks.
    """
    blocks = await _fetch_blocks(db_pool, ha_id, portfolio_id)
    logger.info(f"Doc B: {len(blocks)} blocks for ha_id={ha_id}")

    wb = openpyxl.Workbook()

    # Legend sheet
    _write_legend_sheet(wb)

    # Main sheet
    ws = wb.create_sheet("Building Schedule", 0)
    _write_section_headers(ws)
    _write_question_headers(ws)
    _write_data_rows(ws, blocks)
    _apply_column_widths(ws)
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# DB query — aggregate per block
# ---------------------------------------------------------------------------

async def _fetch_blocks(db_pool, ha_id: str, portfolio_id: Optional[str]) -> list:
    """
    Aggregate property-level data to ONE ROW PER BLOCK.

    The single GROUP BY key is COALESCE(block_reference, address).
    All other fields use aggregate functions so N units collapse into
    one block row correctly.

    LOR and TIV are computed in SQL:
        lor_value           = SUM(sum_insured) * 0.25
        total_insured_value = SUM(sum_insured) * 1.25

    FRA and FRAEW data joined via LATERAL subqueries through silver.blocks.
    """
    sql = """
    -- Block key: use block_reference when present; fall back to address_1 only
    -- (not the full concatenated address which may include city/town suffix).
    WITH block_agg AS (
        SELECT
            COALESCE(p.block_reference, p.address) AS block_reference,
            SUM(p.sum_insured)                              AS sum_insured,
            ROUND((SUM(p.sum_insured) * 0.25)::numeric, 2) AS lor_value,
            ROUND((SUM(p.sum_insured) * 1.25)::numeric, 2) AS total_insured_value,
            COALESCE(SUM(p.units), COUNT(*))                AS units,
            MAX(p.storeys)                                  AS storeys,
            MAX(p.height_max_m)                             AS height_max_m,
            MAX(p.build_year)                               AS build_year,
            MAX(p.wall_construction)                        AS wall_construction,
            MAX(p.floor_construction)                       AS floor_construction,
            MAX(p.roof_construction)                        AS roof_construction,
            BOOL_OR(p.basement)                             AS basement,
            BOOL_OR(p.is_listed)                            AS is_listed
        FROM silver.properties p
        WHERE p.ha_id = $1
        GROUP BY COALESCE(p.block_reference, p.address)
    ),
    -- Most common value per block, with alphabetical tiebreak (ASC) so ties
    -- resolve deterministically and match Ex11 convention (e.g. Rented < Factored).
    block_occ AS (
        -- On a tie, first-inserted unit's occupancy_type wins (matches Ex11 row order).
        SELECT DISTINCT ON (blk)
            blk AS block_reference, occupancy_type
        FROM (
            SELECT
                COALESCE(p.block_reference, p.address) AS blk,
                p.occupancy_type,
                p.property_id,
                COUNT(*) OVER (
                    PARTITION BY COALESCE(p.block_reference, p.address),
                                 p.occupancy_type
                ) AS cnt
            FROM silver.properties p WHERE p.ha_id = $1
        ) sub
        ORDER BY blk, cnt DESC, property_id ASC
    ),
    block_postcode AS (
        SELECT DISTINCT ON (blk)
            blk AS block_reference, postcode
        FROM (
            SELECT
                COALESCE(p.block_reference, p.address) AS blk,
                p.postcode,
                COUNT(*) OVER (
                    PARTITION BY COALESCE(p.block_reference, p.address),
                                 p.postcode
                ) AS cnt
            FROM silver.properties p WHERE p.ha_id = $1
        ) sub
        ORDER BY blk, cnt DESC, postcode ASC
    ),
    block_property_type AS (
        SELECT DISTINCT ON (blk)
            blk AS block_reference, property_type
        FROM (
            SELECT
                COALESCE(p.block_reference, p.address) AS blk,
                p.property_type,
                COUNT(*) OVER (
                    PARTITION BY COALESCE(p.block_reference, p.address),
                                 p.property_type
                ) AS cnt
            FROM silver.properties p WHERE p.ha_id = $1
        ) sub
        ORDER BY blk, cnt DESC, property_type ASC
    ),
    -- Representative address: first unit address in the block
    block_address AS (
        SELECT DISTINCT ON (COALESCE(p.block_reference, p.address))
            COALESCE(p.block_reference, p.address) AS block_reference,
            p.address   AS address,
            p.address_2 AS address_2,
            p.address_3 AS address_3
        FROM silver.properties p
        WHERE p.ha_id = $1
        ORDER BY COALESCE(p.block_reference, p.address), p.property_id
    )
    SELECT
        a.block_reference,
        ad.address,
        ad.address_2,
        ad.address_3,
        pc.postcode,
        a.sum_insured,
        a.lor_value,
        a.total_insured_value,
        a.units,
        a.storeys,
        a.height_max_m,
        a.build_year,
        a.wall_construction,
        a.floor_construction,
        a.roof_construction,
        a.basement,
        a.is_listed,
        o.occupancy_type,
        pt.property_type,
        -- FRA fields (Q32-Q40)
        fra.assessment_date        AS fra_date,
        fra.has_sprinkler_system   AS sprinklers_common,
        fra.has_fire_alarm_system  AS fire_alarms_common,
        fra.has_fire_extinguishers AS other_fire_protection,
        -- FRAEW fields (Q41-Q58)
        fraew.has_combustible_cladding AS combustible_materials,
        fraew.has_remedial_actions     AS remediation_required,
        fraew.interim_measures_detail  AS interim_measures_detail,
        fraew.eps_insulation_present   AS insulation_type,
        fraew.has_combustible_cladding AS cladding_present,
        fraew.has_remedial_actions     AS ews_remediation_required
    FROM block_agg       a
    LEFT JOIN block_occ           o  USING (block_reference)
    LEFT JOIN block_postcode      pc USING (block_reference)
    LEFT JOIN block_property_type pt USING (block_reference)
    LEFT JOIN block_address       ad USING (block_reference)
    -- FRA: latest per block (joined via silver.blocks)
    LEFT JOIN LATERAL (
        SELECT f.*
        FROM silver.fra_features f
        JOIN silver.blocks b ON f.block_id = b.block_id
        WHERE b.ha_id = $1 AND b.name = a.block_reference
        ORDER BY f.assessment_date DESC NULLS LAST
        LIMIT 1
    ) fra ON true
    -- FRAEW: latest per block (joined via silver.blocks)
    LEFT JOIN LATERAL (
        SELECT fw.*
        FROM silver.fraew_features fw
        JOIN silver.blocks b ON fw.block_id = b.block_id
        WHERE b.ha_id = $1 AND b.name = a.block_reference
        ORDER BY fw.assessment_date DESC NULLS LAST
        LIMIT 1
    ) fraew ON true
    ORDER BY a.block_reference
    """
    async with db_pool.acquire() as conn:
        return await conn.fetch(sql, ha_id)


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def _write_section_headers(ws):
    """Row 1: section group headers spanning multiple columns."""
    sections = [
        ("Policy Information",          1,  8,  "insurer"),
        ("General Property Details",    9,  27, "sov"),
        ("Construction",                28, 31, "sov"),
        ("Fire Risk Management",        32, 40, "fra"),
        ("EWS / Cladding Information",  41, 58, "fraew"),
        ("Claim Information",           59, 64, "insurer"),
    ]
    for label, start, end, source in sections:
        cell = ws.cell(row=1, column=start, value=label)
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if end > start:
            ws.merge_cells(
                start_row=1, start_column=start,
                end_row=1, end_column=end
            )


def _write_question_headers(ws):
    """Row 2: individual question headers."""
    for col_idx, (q_num, label, source, _) in enumerate(DOC_B_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=f"{q_num}: {label}")
        cell.font = Font(bold=True, name="Calibri", size=9,
                         color="FFFFFF" if source in ("sov",) else "1F4E79")
        cell.fill = FILLS[source]
        if source == "sov":
            cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 60


def _write_data_rows(ws, blocks):
    for row_idx, block in enumerate(blocks, start=3):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None

        for col_idx, (q_num, label, source, db_field) in enumerate(DOC_B_COLUMNS, start=1):
            if db_field and db_field in block.keys():
                value = block[db_field]
            else:
                value = None

            # Format booleans nicely
            if isinstance(value, bool):
                value = "Yes" if value else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

            # Apply source-based fill colours
            if source in ("fra", "fraew"):
                cell.fill = FILLS[source]
            elif source == "insurer":
                cell.fill = FILLS["insurer"]
            elif fill:
                cell.fill = fill

            # Currency formatting
            if label in ("Sum Insured", "LOR/AA/ICOW", "Total Insured Value") \
                    and value is not None and isinstance(value, (int, float)):
                cell.number_format = '£#,##0.00'

        ws.row_dimensions[row_idx].height = 18


def _apply_column_widths(ws):
    for col_idx in range(1, len(DOC_B_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 60


def _write_legend_sheet(wb):
    ws = wb.create_sheet("Legend")
    ws.sheet_properties.tabColor = "CCCCCC"

    entries = [
        ("White",       FILLS["sov"],     "Populated from SOV upload"),
        ("Light Blue",  FILLS["insurer"], "Completed by insurer / Avid"),
        ("Light Yellow",FILLS["fra"],     "Populated from FRA document"),
        ("Light Orange",FILLS["fraew"],   "Populated from FRAEW document"),
    ]
    ws.cell(row=1, column=1, value="Colour Legend").font = Font(bold=True, size=12)
    for i, (colour, fill, desc) in enumerate(entries, start=3):
        ws.cell(row=i, column=1, value=colour).fill = fill
        ws.cell(row=i, column=2, value=desc)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 45