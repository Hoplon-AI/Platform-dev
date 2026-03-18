"""
doc_a_exporter.py
-----------------
Generates a Doc A (Stock Listing) Excel file from silver.properties.

Doc A = one row per property unit, 35 columns in the Aviva format.
System-filled columns (policy number, GWP etc.) are left blank —
the underwriter or Avid completes those.

Usage:
    from doc_a_exporter import generate_doc_a
    excel_bytes = await generate_doc_a(db_pool, ha_id, ha_name)
"""

import io
import logging
from datetime import date
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Doc A column definitions — in exact order
# ---------------------------------------------------------------------------
# (col_header, db_field_or_None)
# None = insurer-filled, left blank
DOC_A_COLUMNS = [
    ("Client Name",                             "ha_name"),          # injected
    ("Start Date",                              None),               # insurer
    ("End Date",                                None),               # insurer
    ("Policy Reference",                        None),               # insurer
    ("Product Type",                            None),               # insurer
    ("Property Reference",                      "property_reference"),
    ("Block Reference",                         "block_reference"),
    ("Occupancy Type",                          "occupancy_type"),
    ("Deductible",                              None),               # insurer
    ("Flood Deductible",                        None),               # insurer
    ("Storm Deductible",                        None),               # insurer
    ("Basis of Deductible (EEL/SEC)",           None),               # insurer
    ("Address 1",                               "address"),
    ("Address 2",                               "address_2"),
    ("Address 3",                               "address_3"),
    ("Postcode",                                "postcode"),
    ("Number of Units",                         "units"),
    ("Sum Insured",                             "sum_insured"),
    ("Sum Insured Type",                        "sum_insured_type"),
    ("Property Type",                           "property_type"),
    ("Avid Property Type",                      "avid_property_type"),
    ("Wall Construction",                       "wall_construction"),
    ("Roof Construction",                       "roof_construction"),
    ("Floor Construction",                      "floor_construction"),
    ("Year of Build",                           "build_year"),
    ("Age Banding",                             "age_banding"),
    ("Number of Bedrooms",                      "num_bedrooms"),
    ("Number of Storeys",                       "storeys"),
    ("Basement location",                       "basement"),
    ("Listed building (if blank = not listed)", "is_listed"),
    ("Security Features",                       "security_features"),
    ("Fire Protection",                         "fire_protection"),
    ("Alarms",                                  "alarms"),
    ("Flood insured",                           "flood_insured"),
    ("Storm insured",                           "storm_insured"),
    # ── Enrichment columns (from API data) ──
    ("UPRN",                                    "uprn"),
    ("Height (metres)",                         "height_max_m"),
    ("Building Footprint (m²)",                 "building_footprint_m2"),
    ("EPC Rating",                              "epc_rating"),
    ("Listed Grade",                            "listed_grade"),
    ("Enrichment Status",                       "enrichment_status"),
    ("Enrichment Source",                       "enrichment_source"),
]

# Styling
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
INSURER_FILL  = PatternFill("solid", fgColor="D6E4F0")   # light blue - insurer cols
DATA_FILL     = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
NORMAL_FONT   = Font(name="Calibri", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
ALT_ROW_FILL  = PatternFill("solid", fgColor="F5F5F5")

# Insurer-filled column indices (0-based)
INSURER_COLS = {
    DOC_A_COLUMNS.index(c)
    for c in DOC_A_COLUMNS
    if c[1] is None and c[0] not in ("Client Name",)
}


# ---------------------------------------------------------------------------
# Public function
# ---------------------------------------------------------------------------

async def generate_doc_a(
    db_pool,
    ha_id: str,
    ha_name: str,
    portfolio_id: Optional[str] = None,
) -> bytes:
    """
    Query silver.properties for the given HA and generate Doc A Excel bytes.
    Matches the template structure: Stock Listing, Validation pivot, Notes,
    Client notes, Insured Type.
    """
    rows = await _fetch_properties(db_pool, ha_id, portfolio_id)
    logger.info(f"Doc A: {len(rows)} properties for ha_id={ha_id}")

    wb = openpyxl.Workbook()

    # --- Sheet 1: Stock Listing ---
    ws = wb.active
    ws.title = "Stock Listing"
    _write_header_row(ws, ha_name)
    _write_data_rows(ws, rows, ha_name)
    _apply_column_widths(ws)
    _freeze_header(ws)

    # --- Sheet 2: Validation pivot ---
    ws_pivot = wb.create_sheet("Validation pivot")
    _write_validation_pivot(ws_pivot, rows)

    # --- Sheet 3: Notes ---
    ws_notes = wb.create_sheet("Notes")
    _write_notes_sheet(ws_notes)

    # --- Sheet 4: Client notes ---
    wb.create_sheet("Client notes")

    # --- Sheet 5: Insured Type ---
    ws_ins = wb.create_sheet("Insured Type")
    _write_insured_type_sheet(ws_ins)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# DB query
# ---------------------------------------------------------------------------

async def _fetch_properties(db_pool, ha_id: str, portfolio_id: Optional[str]) -> list:
    sql = """
    SELECT
        p.property_reference,
        p.block_reference,
        p.occupancy_type,
        p.address,
        p.address_2,
        p.address_3,
        p.postcode,
        p.units,
        p.sum_insured,
        p.sum_insured_type,
        p.property_type,
        p.avid_property_type,
        p.wall_construction,
        p.roof_construction,
        p.floor_construction,
        p.build_year,
        p.age_banding,
        p.num_bedrooms,
        p.storeys,
        p.basement,
        p.is_listed,
        p.security_features,
        p.fire_protection,
        p.alarms,
        p.flood_insured,
        p.storm_insured,
        p.uprn,
        p.height_max_m,
        p.building_footprint_m2,
        p.epc_rating,
        p.listed_grade,
        p.enrichment_status,
        p.enrichment_source
    FROM silver.properties p
    WHERE p.ha_id = $1
    ORDER BY p.block_reference NULLS LAST, p.address
    """
    async with db_pool.acquire() as conn:
        return await conn.fetch(sql, ha_id)


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def _write_header_row(ws, ha_name: str):
    for col_idx, (header, _) in enumerate(DOC_A_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = INSURER_FILL if (col_idx - 1) in INSURER_COLS else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 40


def _write_data_rows(ws, rows, ha_name: str):
    for row_idx, db_row in enumerate(rows, start=2):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else DATA_FILL

        for col_idx, (header, db_field) in enumerate(DOC_A_COLUMNS, start=1):
            if db_field is None:
                value = None   # insurer fills this
            elif db_field == "ha_name":
                value = ha_name
            else:
                value = db_row[db_field] if db_field in db_row.keys() else None

            # Format booleans nicely
            if isinstance(value, bool):
                value = "Yes" if value else "No"

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NORMAL_FONT
            cell.fill = INSURER_FILL if (col_idx - 1) in INSURER_COLS else fill
            cell.alignment = Alignment(vertical="center")
            cell.border = THIN_BORDER

            # Currency formatting for sum_insured
            if header == "Sum Insured" and value is not None:
                cell.number_format = '£#,##0.00'


def _apply_column_widths(ws):
    widths = {
        1: 25,   # Client Name
        6: 15,   # Property Reference
        7: 12,   # Block Reference
        8: 15,   # Occupancy Type
        13: 40,  # Address 1
        14: 25,  # Address 2
        15: 20,  # Address 3
        16: 12,  # Postcode
        17: 8,   # Units
        18: 14,  # Sum Insured
        19: 30,  # Sum Insured Type
        20: 20,  # Property Type
        22: 20,  # Wall Construction
        23: 20,  # Roof Construction
        24: 20,  # Floor Construction
        25: 10,  # Year of Build
        26: 15,  # Age Banding
    }
    total_cols = len(DOC_A_COLUMNS)
    for col in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 12)


def _freeze_header(ws):
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Validation Pivot sheet
# ---------------------------------------------------------------------------

PIVOT_HEADER_FILL  = PatternFill("solid", fgColor="2E75B6")
PIVOT_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
PIVOT_SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
PIVOT_TOTAL_FONT   = Font(bold=True, name="Calibri", size=10)
PIVOT_NORMAL_FONT  = Font(name="Calibri", size=10)
GBP_FORMAT         = '£#,##0.00'
PCT_FORMAT         = '0.0%'


def _write_validation_pivot(ws, rows):
    """
    Compute and write four pivot tables side-by-side, matching the template:
      A-C:  By Tenancy/Ownership  (occupancy_type)
      E-J:  By Block              (block_reference, includes LOR/AA, TIV, >£5m)
      L-O:  By Property Type      (avid_property_type)
      Q-T:  By Age Banding        (age_banding)
    """
    from collections import defaultdict

    # Aggregate
    tenancy   = defaultdict(lambda: {"units": 0, "si": 0.0})
    block     = defaultdict(lambda: {"units": 0, "si": 0.0})
    prop_type = defaultdict(lambda: {"units": 0, "si": 0.0})
    age       = defaultdict(lambda: {"units": 0, "si": 0.0})

    total_units = 0
    total_si    = 0.0

    for r in rows:
        u  = r["units"] or 1
        si = float(r["sum_insured"] or 0)
        total_units += u
        total_si    += si
        tenancy  [r["occupancy_type"]     or "(blank)"]["units"] += u
        tenancy  [r["occupancy_type"]     or "(blank)"]["si"]    += si
        block    [r["block_reference"]    or "(blank)"]["units"] += u
        block    [r["block_reference"]    or "(blank)"]["si"]    += si
        prop_type[r["avid_property_type"] or "(blank)"]["units"] += u
        prop_type[r["avid_property_type"] or "(blank)"]["si"]    += si
        age      [r["age_banding"]        or "(blank)"]["units"] += u
        age      [r["age_banding"]        or "(blank)"]["si"]    += si

    LOR_RATE = 0.25  # standard LOR/AA assumption

    def _section_header(ws, row, col, title, subheaders):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = PIVOT_HEADER_FONT
        cell.fill = PIVOT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + len(subheaders) - 1)
        for i, h in enumerate(subheaders):
            c = ws.cell(row=row + 1, column=col + i, value=h)
            c.font = PIVOT_HEADER_FONT
            c.fill = PIVOT_SECTION_FILL
            c.alignment = Alignment(horizontal="center", wrap_text=True)

    def _total_row(ws, row, col, label, units, si, extra_cols=None):
        vals = [label, units, si] + (extra_cols or [])
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=col + i, value=v)
            c.font = PIVOT_TOTAL_FONT
            if i == 2:
                c.number_format = GBP_FORMAT

    def _data_row(ws, row, col, label, units, si, extra_cols=None):
        vals = [label, units, si] + (extra_cols or [])
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=col + i, value=v)
            c.font = PIVOT_NORMAL_FONT
            if i == 2:
                c.number_format = GBP_FORMAT
            if extra_cols and i >= 3:
                # percentage columns
                if isinstance(v, float) and 0 <= v <= 1:
                    c.number_format = PCT_FORMAT

    # --- Section 1: By Tenancy (col A=1) ---
    _section_header(ws, 1, 1, "By Tenancy/Ownership",
                    ["Row Labels", "Count of Units", "Sum of Sum Insured"])
    r = 3
    for key in sorted(tenancy):
        _data_row(ws, r, 1, key, tenancy[key]["units"], tenancy[key]["si"])
        r += 1
    _total_row(ws, r, 1, "Grand Total", total_units, total_si)

    # --- Section 2: By Block (col E=5) ---
    _section_header(ws, 1, 5, "By Block (Block Reference)",
                    ["Row Labels", "Count of Units", "Sum of Sum Insured",
                     "LOR/AA", "TIV", "Over £5m?"])
    r = 3
    for key in sorted(block):
        si   = block[key]["si"]
        lor  = si * LOR_RATE
        tiv  = si + lor
        flag = "Yes" if tiv > 5_000_000 else "No"
        _data_row(ws, r, 5, key, block[key]["units"], si, [lor, tiv, flag])
        r += 1
    _total_row(ws, r, 5, "Grand Total", total_units, total_si,
               [total_si * LOR_RATE, total_si * (1 + LOR_RATE), ""])

    # --- Section 3: By Property Type (col L=12) ---
    _section_header(ws, 1, 12, "Prop Type",
                    ["Row Labels", "Count of Units", "Sum of Sum Insured", "% of Total"])
    r = 3
    for key in sorted(prop_type):
        si  = prop_type[key]["si"]
        pct = si / total_si if total_si else 0
        _data_row(ws, r, 12, key, prop_type[key]["units"], si, [pct])
        r += 1
    _total_row(ws, r, 12, "Grand Total", total_units, total_si, [1.0])

    # --- Section 4: By Age Banding (col Q=17) ---
    _section_header(ws, 1, 17, "Age",
                    ["Row Labels", "Count of Units", "Sum of Sum Insured", "% of Total"])
    r = 3
    for key in sorted(age):
        si  = age[key]["si"]
        pct = si / total_si if total_si else 0
        _data_row(ws, r, 17, key, age[key]["units"], si, [pct])
        r += 1
    _total_row(ws, r, 17, "Grand Total", total_units, total_si, [1.0])

    # Column widths
    for col, w in [(1, 20), (2, 12), (3, 16), (5, 20), (6, 12), (7, 16),
                   (8, 14), (9, 16), (10, 12), (12, 20), (13, 12), (14, 16),
                   (15, 12), (17, 20), (18, 12), (19, 16), (20, 12)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Notes sheet
# ---------------------------------------------------------------------------

def _write_notes_sheet(ws):
    NOTES_CONTENT = [
        (4,  "Notes"),
        (6,  "All information to be included where available"),
        (7,  "It is recognised that not all fields will be able to be populated where clients "
             "have not provided information however will look to include where available"),
        (8,  "Best endeavours to be used to fill in missing info (e.g. if no property type "
             "given but address has flat in prefix then can note as flats)"),
        (10, None),  # spacer
    ]
    FIELD_NOTES = [
        ("Field",              "Comment"),
        ("Client Name",        ""),
        ("Start Date",         ""),
        ("End Date",           ""),
        ("Policy Reference",   ""),
        ("Product Type",       "Housing Association/Leasehold"),
        ("Property Reference", "If provided by client in order to aide in cross referencing back to raw listing"),
        ("Occupancy Type",     "Condensed list to aide in MI based on below matrix"),
        ("Deductible",         "Excess based on occupancy type (e.g. Rented or leasehold)"),
        ("Deductible type",    "If excess is SEC or EEL (recognising rented/LH split above)"),
        ("Address fields",     "Address as provided by clients"),
        ("Units",              "Number of units for line in question"),
        ("Sum Insured",        "Sum insured for unit"),
        ("Sum Insured Type",   "E.g. declared by client or per unit average"),
        ("Property Type",      "As provided by client"),
        ("Avid Property Type", "Condensed list to aide in MI based on below matrix"),
        ("Wall Construction",  ""),
        ("Roof Construction",  ""),
        ("Floor Construction", ""),
        ("Property Age",       ""),
        ("Age Banding",        "Condensed list to aide in MI based on below matrix"),
        ("Number of Storeys",  "Number of storeys in block"),
        ("Security Features",  ""),
        ("Fire Protection",    ""),
        ("Alarms",             ""),
        ("Block Reference",    "If a block is known then to add onto reference to add in agg reporting"),
    ]

    header_font = Font(bold=True, name="Calibri", size=10)
    normal_font = Font(name="Calibri", size=10)

    for row_num, text in NOTES_CONTENT:
        if text:
            c = ws.cell(row=row_num, column=1, value=text)
            c.font = header_font if row_num == 4 else normal_font

    start_row = 10
    for i, (field, comment) in enumerate(FIELD_NOTES):
        r = start_row + i
        c1 = ws.cell(row=r, column=1, value=field)
        c2 = ws.cell(row=r, column=2, value=comment)
        if i == 0:
            c1.font = header_font
            c2.font = header_font
        else:
            c1.font = normal_font
            c2.font = normal_font

    # Lookup matrix rows
    matrix_row = start_row + len(FIELD_NOTES) + 2
    matrices = [
        ("Property Type",      ["House/Bungalow", "Flat", "Commercial", "Hostel",
                                 "Garage", "Other", "Not Known",
                                 "Commercial contents", "All Risks",
                                 "Business Interruption", "Landlords Contents", "Money"]),
        ("Age Banding",        ["Pre 1900", "1901-1919", "1920-1944",
                                 "1945-1979", "1980-2000", "2001+"]),
        ("Occupancy Type",     ["Rented", "Leasehold", "Commercial", "Factored",
                                 "Other", "Commercial contents",
                                 "All Risks", "Business Interruption",
                                 "Landlords Contents", "Money"]),
    ]
    for col_offset, (label, values) in enumerate(matrices):
        col = 1 + col_offset
        c = ws.cell(row=matrix_row, column=col, value=label)
        c.font = header_font
        for j, v in enumerate(values):
            ws.cell(row=matrix_row + 1 + j, column=col, value=v).font = normal_font

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


# ---------------------------------------------------------------------------
# Insured Type sheet
# ---------------------------------------------------------------------------

def _write_insured_type_sheet(ws):
    c = ws.cell(row=4, column=1, value="Insured Type")
    c.font = Font(bold=True, name="Calibri", size=10)
    ws.column_dimensions["A"].width = 20